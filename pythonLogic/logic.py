"""
=============================================================================
  LPG AUTO DISPATCH OPTIMIZER  –  Full Standalone Script
  Run:  python3 FULL_LPG_OPTIMIZATION.py
  Output: LPG_DISPATCH_PLAN_<date>.xlsx  (same folder as this script)

  PERSISTENT STATE:
    truck_positions.json  – saved after every run, loaded at start of next run
    • First run ever   → trucks placed by max-dispersion across all 81 stations
    • Every later run  → trucks start exactly where they parked last time
    • Reset positions  → delete truck_positions.json and re-run
=============================================================================
"""

import pandas as pd
import requests
import math
import time
import os
import sys
import json
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

STATIONS_FILE      = "clean_stationss.xlsx"
SOURCES_FILE       = "sources.xlsx"
POSITIONS_FILE     = "truck_positions.json"    # persistent state file
GOOGLE_API_KEY     = "AIzaSyA8oVRSa2W2IzX9hg4vnaKM6hwkGRGnsP4"

FLEET = [
    {"type": "12MT", "count": 23, "capacity_mt": 12, "capacity_lt": 12 * 1810},
    {"type": "7MT",  "count":  7, "capacity_mt":  7, "capacity_lt":  7 * 1810},
]

MT_TO_LITERS        = 1810
TRANSPORT_FLAT      = 1750      # Rs/MT for dist < 100 km
TRANSPORT_PER_KM    = 6.8       # Rs/MT/RTKM for dist >= 100 km
MAX_STOPS_PER_TRUCK = 3
MAX_GROUPING_KM     = 80

# ═══════════════════════════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════

def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0
    r = math.radians
    dlat = r(lat2 - lat1); dlon = r(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(r(lat1)) * math.cos(r(lat2)) * math.sin(dlon/2)**2
    return 2 * R * math.asin(math.sqrt(a))

def parse_coords(s):
    parts = str(s).split(',')
    return float(parts[0].strip()), float(parts[1].strip())

_route_cache = {}

def get_road_info(olat, olon, dlat, dlon):
    key = (round(olat, 4), round(olon, 4), round(dlat, 4), round(dlon, 4))
    if key in _route_cache:
        return _route_cache[key]
    try:
        url = "https://routes.googleapis.com/directions/v2:computeRoutes"
        headers = {
            "Content-Type": "application/json",
            "X-Goog-Api-Key": GOOGLE_API_KEY,
            "X-Goog-FieldMask": "routes.distanceMeters,routes.travelAdvisory.tollInfo",
        }
        body = {
            "origin":      {"location": {"latLng": {"latitude": olat, "longitude": olon}}},
            "destination": {"location": {"latLng": {"latitude": dlat, "longitude": dlon}}},
            "travelMode":  "DRIVE",
            "extraComputations": ["TOLLS"],
            "routeModifiers": {"vehicleInfo": {"emissionType": "DIESEL"}},
        }
        resp = requests.post(url, headers=headers, json=body, timeout=12)
        data = resp.json()
        if "routes" in data and data["routes"]:
            route   = data["routes"][0]
            dist_km = route["distanceMeters"] / 1000.0
            toll    = 0.0
            for p in route.get("travelAdvisory", {}).get("tollInfo", {}).get("estimatedPrice", []):
                toll += float(p.get("units", 0)) + float(p.get("nanos", 0)) / 1e9
            _route_cache[key] = (dist_km, toll)
            return dist_km, toll
    except Exception:
        pass
    dist = haversine(olat, olon, dlat, dlon) * 1.3
    _route_cache[key] = (dist, 0.0)
    return dist, 0.0

def calc_transport_cost(dist_km, qty_mt):
    if dist_km < 100:
        return TRANSPORT_FLAT * qty_mt
    return TRANSPORT_PER_KM * qty_mt * dist_km * 2

# ═══════════════════════════════════════════════════════════════════════════════
#  DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════════

def resolve_path(filename):
    base = os.path.dirname(os.path.abspath(__file__))
    p = os.path.join(base, filename)
    if os.path.exists(p):
        return p
    fallback = f"/mnt/user-data/uploads/{filename}"
    if os.path.exists(fallback):
        return fallback
    raise FileNotFoundError(f"Cannot find {filename}")

def load_data():
    stations = pd.read_excel(resolve_path(STATIONS_FILE))
    sources  = pd.read_excel(resolve_path(SOURCES_FILE))
    stations.columns = [c.strip() for c in stations.columns]
    sources.columns  = [c.strip() for c in sources.columns]
    if "Stations " in stations.columns:
        stations.rename(columns={"Stations ": "Stations"}, inplace=True)
    if "Source_ID " in sources.columns:
        sources.rename(columns={"Source_ID ": "Source_ID"}, inplace=True)
    stations["Stations"] = stations["Stations"].str.strip()
    stations["lat"], stations["lon"] = zip(*stations["Coordinates"].map(parse_coords))
    sources["lat"],  sources["lon"]  = zip(*sources["Coordinates"].map(parse_coords))
    return stations, sources

# ═══════════════════════════════════════════════════════════════════════════════
#  PERSISTENT TRUCK POSITIONS
# ═══════════════════════════════════════════════════════════════════════════════

def positions_path():
    """
    Store the JSON file alongside the outputs Excel so it persists.
    When running locally (no /mnt/user-data/outputs), store beside the script.
    """
    outputs = "/mnt/user-data/outputs"
    if os.path.isdir(outputs):
        return os.path.join(outputs, POSITIONS_FILE)
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, POSITIONS_FILE)

def load_saved_positions():
    """Returns dict {truck_id: {station, lat, lon, type}} or None if no file."""
    p = positions_path()
    if os.path.exists(p):
        with open(p, "r") as f:
            data = json.load(f)
        print(f"      ✓ Loaded saved positions from: {p}")
        return data
    print(f"      ℹ  No saved positions file found – this is the first run.")
    return None

def save_positions(trucks):
    """Saves every truck's current parked position to JSON."""
    data = {
        t["truck_id"]: {
            "station": t["parked_station"],
            "lat":     t["parked_lat"],
            "lon":     t["parked_lon"],
            "type":    t["type"],
        }
        for t in trucks
    }
    p = positions_path()
    with open(p, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"      ✓ Positions saved → {p}")
    print(f"        (Next run will load from this file)")

# ═══════════════════════════════════════════════════════════════════════════════
#  FIRST-RUN PLACEMENT  (greedy max-dispersion)
# ═══════════════════════════════════════════════════════════════════════════════

def dispersion_indices(stations, n):
    coords = list(zip(stations["lat"], stations["lon"]))
    chosen = [0]
    while len(chosen) < n:
        best_i, best_d = -1, -1.0
        for i in range(len(coords)):
            if i in chosen:
                continue
            md = min(haversine(coords[i][0], coords[i][1],
                               coords[c][0], coords[c][1]) for c in chosen)
            if md > best_d:
                best_d, best_i = md, i
        chosen.append(best_i)
    return chosen

# ═══════════════════════════════════════════════════════════════════════════════
#  BUILD FLEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_fleet(stations):
    """
    Returns list of truck dicts with parked_station/lat/lon populated.
    Uses saved JSON if available, otherwise does first-run dispersion.
    """
    saved = load_saved_positions()

    # Build truck list (positions filled below)
    trucks = []
    num = 1
    for ft in FLEET:
        for _ in range(ft["count"]):
            trucks.append({
                "truck_id":    f"T{num:02d}",
                "type":        ft["type"],
                "capacity_mt": ft["capacity_mt"],
                "capacity_lt": ft["capacity_lt"],
                "parked_station": None,
                "parked_lat":     None,
                "parked_lon":     None,
            })
            num += 1

    if saved:
        # ── Restore last-known positions ──────────────────────────────────────
        for t in trucks:
            if t["truck_id"] in saved:
                pos = saved[t["truck_id"]]
                t["parked_station"] = pos["station"]
                t["parked_lat"]     = pos["lat"]
                t["parked_lon"]     = pos["lon"]
            else:
                # Truck ID not in saved file (fleet expanded) – assign a new spot
                coords = list(zip(stations["lat"], stations["lon"]))
                used   = {(round(v["lat"],4), round(v["lon"],4)) for v in saved.values()}
                best_i, best_d = 0, -1.0
                for i,(la,lo) in enumerate(coords):
                    d = min(haversine(la,lo,u[0],u[1]) for u in used) if used else 999.0
                    if d > best_d:
                        best_d, best_i = d, i
                row = stations.iloc[best_i]
                t["parked_station"] = row["Stations"]
                t["parked_lat"]     = float(row["lat"])
                t["parked_lon"]     = float(row["lon"])
                used.add((round(row["lat"],4), round(row["lon"],4)))

        print("      ↩  Trucks restored to their positions from the previous run")

    else:
        # ── First run: spread trucks across Tamil Nadu ────────────────────────
        indices = dispersion_indices(stations, len(trucks))
        for t, idx in zip(trucks, indices):
            row = stations.iloc[idx]
            t["parked_station"] = row["Stations"]
            t["parked_lat"]     = float(row["lat"])
            t["parked_lon"]     = float(row["lon"])
        print(f"      🆕 First run – trucks spread across all {len(stations)} stations")

    return trucks

# ═══════════════════════════════════════════════════════════════════════════════
#  BALANCED RUN PARTITIONING
# ═══════════════════════════════════════════════════════════════════════════════

def balanced_partition(stations_in_group, max_stops, max_grouping_km):
    """
    Split a list of stations (all sharing the same source) into delivery runs
    so that truck loads are as balanced as possible.

    Rules
    -----
    • Each run holds at most max_stops stations.
    • Stations in the same run must be within max_grouping_km of each other
      (nearest-neighbour chain check).
    • When N stations exist and N is exactly divisible by k (where k < max_stops),
      prefer k-stop runs over (max_stops)-stop + remainder runs.
      Example: 4 stations → 2 runs of 2  (not 1 run of 3 + 1 run of 1)
               6 stations → 2 runs of 3  (balanced, no change needed)
               5 stations → 1 run of 3 + 1 run of 2  (unavoidable)
    • Within each run the order is nearest-neighbour starting from the station
      closest to the source.
    """
    if not stations_in_group:
        return []

    n = len(stations_in_group)

    # ── Determine ideal run size ──────────────────────────────────────────────
    # Find the smallest run_size in [2 .. max_stops] such that n % run_size == 0
    # (perfectly balanced).  If none divides evenly, fall back to max_stops.
    ideal_size = max_stops
    for size in range(2, max_stops + 1):
        if n % size == 0:
            ideal_size = size
            break   # take the smallest that divides evenly → most balanced

    # ── Build runs using nearest-neighbour, respecting ideal_size ────────────
    remaining = stations_in_group[:]   # already sorted by dist-from-source
    runs = []

    while remaining:
        run = [remaining.pop(0)]
        target = ideal_size if len(remaining) + 1 >= ideal_size else max_stops

        while len(run) < target and remaining:
            last = run[-1]
            ni = min(
                range(len(remaining)),
                key=lambda i: haversine(
                    last["station_lat"], last["station_lon"],
                    remaining[i]["station_lat"], remaining[i]["station_lon"]
                )
            )
            d_next = haversine(
                last["station_lat"], last["station_lon"],
                remaining[ni]["station_lat"], remaining[ni]["station_lon"]
            )
            if d_next <= max_grouping_km:
                run.append(remaining.pop(ni))
            else:
                break   # next nearest is too far – close this run

        # ── Rebalance: if this run hit max_stops but leaving just 1 behind,
        #    split evenly (e.g. 4 stations that weren't evenly divisible above
        #    but the proximity check forced a 3+1 → convert to 2+2) ──────────
        if len(run) == max_stops and len(remaining) == 1:
            # Give the last station of this run back to remaining
            leftover = run.pop()
            remaining.insert(0, leftover)
            # remaining now has 2 items → they'll form the next run of 2

        runs.append(run)

    return runs


# ═══════════════════════════════════════════════════════════════════════════════
#  CORE OPTIMIZATION
# ═══════════════════════════════════════════════════════════════════════════════

def run_optimization():
    print("=" * 65)
    print("  LPG AUTO DISPATCH OPTIMIZER")
    print(f"  Date : {datetime.today().strftime('%d-%m-%Y  %H:%M')}")
    print("=" * 65)

    # 1. Load data
    print("\n[1/6] Loading station & source data …")
    stations, sources = load_data()
    print(f"      Stations : {len(stations)}   |   Sources : {len(sources)}")

    # 2. Stations needing LPG
    needing = stations[stations["Now"].str.strip().str.upper() == "NO"].copy()
    print(f"\n[2/6] Stations needing LPG today : {len(needing)}")
    if needing.empty:
        print("      ✓ No deliveries needed today.")
        return None, None, None
    for _, r in needing.iterrows():
        print(f"      – {r['Stations']}  ({r['Usable Lt']:,.0f} Lt)")

    # 3. Build fleet – load persistent positions or first-run dispersion
    print("\n[3/6] Loading truck positions …")
    trucks = build_fleet(stations)

    # ── Snapshot: where every truck is at the START of this run ──────────────
    # This is frozen here and never changes – it is what we write to the Excel
    # as "Initial Parked At (Start of This Run)" for every truck.
    start_positions = {
        t["truck_id"]: {
            "station": t["parked_station"],
            "lat":     t["parked_lat"],
            "lon":     t["parked_lon"],
        }
        for t in trucks
    }

    for t in trucks[:5]:
        print(f"      {t['truck_id']} ({t['type']}) → {t['parked_station']}")
    if len(trucks) > 5:
        print(f"      … and {len(trucks)-5} more trucks")

    # 4. Best source per needing station
    print("\n[4/6] Finding best source per station (Google Routes API) …")
    station_data = []
    for _, srow in needing.iterrows():
        slat, slon = float(srow["lat"]), float(srow["lon"])
        sname      = srow["Stations"]
        needed_lt  = float(srow["Usable Lt"])
        needed_mt  = needed_lt / MT_TO_LITERS

        best_src, best_total = None, float("inf")
        best_dist = best_toll = best_tc = best_pc = None

        for _, src in sources.iterrows():
            dist_km, toll = get_road_info(float(src["lat"]), float(src["lon"]), slat, slon)
            tc  = calc_transport_cost(dist_km, needed_mt)
            pc  = float(src["Price / MT Ex Terminal"]) * needed_mt
            tot = pc + tc + toll
            if tot < best_total:
                best_total = tot
                best_src   = src.copy()
                best_dist, best_toll, best_tc, best_pc = dist_km, toll, tc, pc
            time.sleep(0.04)

        station_data.append({
            "station":        sname,
            "station_lat":    slat,
            "station_lon":    slon,
            "needed_lt":      needed_lt,
            "needed_mt":      needed_mt,
            "source_id":      best_src["Source_ID"],
            "source_name":    best_src["Source_Name"],
            "source_lat":     float(best_src["lat"]),
            "source_lon":     float(best_src["lon"]),
            "price_mt":       float(best_src["Price / MT Ex Terminal"]),
            "dist_km":        best_dist,
            "toll_cost":      best_toll,
            "transport_cost": best_tc,
            "purchase_cost":  best_pc,
            "total_cost":     best_total,
        })
        print(f"      ✓ {sname[:45]:<45} → {best_src['Source_ID']}  (₹{best_total:,.0f})")

    # 5. Group into delivery runs, assign trucks
    print(f"\n[5/6] Grouping into delivery runs (max {MAX_STOPS_PER_TRUCK} stops/truck) …")

    by_source = defaultdict(list)
    for sd in station_data:
        by_source[sd["source_id"]].append(sd)

    delivery_plans  = []
    truck_available = {t["truck_id"]: True for t in trucks}
    truck_by_id     = {t["truck_id"]: t    for t in trucks}

    for src_id, sds in by_source.items():
        src_lat = sds[0]["source_lat"]
        src_lon = sds[0]["source_lon"]

        for sd in sds:
            sd["_d_src"] = haversine(src_lat, src_lon, sd["station_lat"], sd["station_lon"])
        sds.sort(key=lambda x: x["_d_src"])

        runs = balanced_partition(sds, MAX_STOPS_PER_TRUCK, MAX_GROUPING_KM)

        for run in runs:
            total_lt = sum(r["needed_lt"] for r in run)
            total_mt = total_lt / MT_TO_LITERS

            candidates = [(haversine(t["parked_lat"], t["parked_lon"], src_lat, src_lon), t)
                          for t in trucks
                          if truck_available[t["truck_id"]] and t["capacity_mt"] >= total_mt]
            if not candidates:
                candidates = [(haversine(t["parked_lat"], t["parked_lon"], src_lat, src_lon), t)
                              for t in trucks if truck_available[t["truck_id"]]]
            if not candidates:
                candidates = [(haversine(t["parked_lat"], t["parked_lon"], src_lat, src_lon), t)
                              for t in trucks]
            candidates.sort(key=lambda x: x[0])
            _, chosen = candidates[0]
            truck_available[chosen["truck_id"]] = False

            tk_src_dist, tk_src_toll = get_road_info(
                chosen["parked_lat"], chosen["parked_lon"], src_lat, src_lon)
            time.sleep(0.04)

            stops_detail = []
            prev_lat, prev_lon = src_lat, src_lon
            for stop in run:
                d, tl = get_road_info(prev_lat, prev_lon,
                                      stop["station_lat"], stop["station_lon"])
                stops_detail.append({
                    "station":        stop["station"],
                    "needed_lt":      stop["needed_lt"],
                    "needed_mt":      stop["needed_mt"],
                    "dist_km":        round(d, 1),
                    "toll":           round(tl, 2),
                    "transport_cost": round(stop["transport_cost"], 2),
                    "purchase_cost":  round(stop["purchase_cost"], 2),
                    "total_cost":     round(stop["total_cost"], 2),
                    "station_lat":    stop["station_lat"],
                    "station_lon":    stop["station_lon"],
                })
                prev_lat, prev_lon = stop["station_lat"], stop["station_lon"]
                time.sleep(0.04)

            final_park = stops_detail[-1]["station"]
            final_lat  = stops_detail[-1]["station_lat"]
            final_lon  = stops_detail[-1]["station_lon"]

            tot_purchase  = round(sum(s["purchase_cost"]  for s in stops_detail), 2)
            tot_transport = round(sum(s["transport_cost"] for s in stops_detail), 2)
            tot_toll      = round(sum(s["toll"] for s in stops_detail) + tk_src_toll, 2)
            grand_total   = round(tot_purchase + tot_transport + tot_toll, 2)

            delivery_plans.append({
                "truck_id":         chosen["truck_id"],
                "truck_type":       chosen["type"],
                "capacity_lt":      chosen["capacity_lt"],
                # ← position loaded from JSON (or first-run dispersion)
                "initial_park":     start_positions[chosen["truck_id"]]["station"],
                "initial_park_lat": start_positions[chosen["truck_id"]]["lat"],
                "initial_park_lon": start_positions[chosen["truck_id"]]["lon"],
                "source_id":        src_id,
                "source_name":      run[0]["source_name"],
                "source_lat":       src_lat,
                "source_lon":       src_lon,
                "tk_src_dist":      round(tk_src_dist, 1),
                "tk_src_toll":      round(tk_src_toll, 2),
                "stops":            stops_detail,
                "final_park":       final_park,
                "final_lat":        final_lat,
                "final_lon":        final_lon,
                "total_lt":         round(total_lt),
                "total_mt":         round(total_mt, 3),
                "tot_purchase":     tot_purchase,
                "tot_transport":    tot_transport,
                "tot_toll":         tot_toll,
                "grand_total":      grand_total,
            })

            # ── Mutate truck object so positions are updated for JSON save ────
            truck_by_id[chosen["truck_id"]]["parked_station"] = final_park
            truck_by_id[chosen["truck_id"]]["parked_lat"]     = final_lat
            truck_by_id[chosen["truck_id"]]["parked_lon"]     = final_lon

            print(f"      {chosen['truck_id']} ({chosen['type']}) | {src_id} | "
                  f"[starts: {start_positions[chosen['truck_id']]['station']}] → "
                  f"{' → '.join(s['station'] for s in stops_detail)}")

    print(f"\n      Total runs planned : {len(delivery_plans)}")

    # 6. Fleet status table
    used_ids = {dp["truck_id"] for dp in delivery_plans}
    fleet_status = []
    for t in trucks:
        fleet_status.append({
            "truck_id":     t["truck_id"],
            "type":         t["type"],
            "status":       "DEPLOYED" if t["truck_id"] in used_ids else "STANDBY",
            # start_positions = where they were at the beginning of THIS run
            "initial_park": start_positions[t["truck_id"]]["station"],
            # truck_by_id has updated position for deployed trucks, same for standby
            "final_park":   truck_by_id[t["truck_id"]]["parked_station"],
        })

    # ── Save updated positions → next run will load these ─────────────────────
    print("\n      Saving updated truck positions for next run …")
    save_positions(trucks)

    return delivery_plans, fleet_status, trucks

# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

C_BLUE_DARK    = "1F4E79";  C_BLUE_LIGHT  = "D6E4F0";  C_BLUE_ALT   = "EBF3FB"
C_GREEN_DARK   = "375623";  C_GREEN_LIGHT = "E2EFDA"
C_ORANGE_DARK  = "C55A11";  C_ORANGE_LIGHT= "FCE4D6"
C_PURPLE_DARK  = "6B2C91";  C_YELLOW      = "FFF2CC"
C_GREY_LIGHT   = "F2F2F2";  C_WHITE       = "FFFFFF"
C_DEPLOYED     = "C6EFCE";  C_STANDBY     = "FFEB9C"

def _fill(h):  return PatternFill("solid", start_color=h)
def _font(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)
def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _hdr(ws, row, col, val, fh=C_BLUE_DARK):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
    c.fill = _fill(fh); c.border = _border(); c.alignment = _center()

def _cell(ws, row, col, val, fmt=None, fh=None, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _font(bold=bold); c.border = _border(); c.alignment = _left()
    if fmt: c.number_format = fmt
    if fh:  c.fill = _fill(fh)

def _title(ws, rng, text, fh=C_BLUE_DARK):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = text
    c.font  = Font(name="Arial", bold=True, size=14, color=C_WHITE)
    c.fill  = _fill(fh); c.alignment = _center()
    ws.row_dimensions[1].height = 32

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_excel(delivery_plans, fleet_status, trucks, output_path):
    wb        = Workbook()
    today_str = datetime.today().strftime("%d-%m-%Y")
    MF, NF    = "#,##0.00", "#,##0"

    # ── Sheet 1 : Delivery Plan ───────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Delivery Plan"; ws1.freeze_panes = "A3"
    _title(ws1, "A1:R1", f"LPG AUTO DISPATCH – DAILY DELIVERY PLAN   ({today_str})")
    h1 = ["Truck\nID","Truck\nType","Capacity\n(Lt)","Starts From\n(Initial Park)",
          "Goes To\nSource","Source Name","Truck→Source\n(km)","Truck→Source\nToll (₹)",
          "Stop 1","Stop 2","Stop 3",
          "Total Delivered\n(Lt)","Total Delivered\n(MT)",
          "Purchase\nCost (₹)","Transport\nCost (₹)","Toll\nCost (₹)",
          "Grand Total\nCost (₹)","Parked At\n(End of Day)"]
    for ci,h in enumerate(h1,1): _hdr(ws1, 2, ci, h)
    ws1.row_dimensions[2].height = 44

    for ri, dp in enumerate(delivery_plans, 3):
        alt = C_BLUE_ALT if ri%2==0 else C_WHITE
        sn  = [s["station"] for s in dp["stops"]]
        while len(sn)<3: sn.append("")
        vals = [dp["truck_id"],dp["truck_type"],dp["capacity_lt"],dp["initial_park"],
                dp["source_id"],dp["source_name"],dp["tk_src_dist"],dp["tk_src_toll"],
                sn[0],sn[1],sn[2],dp["total_lt"],dp["total_mt"],
                dp["tot_purchase"],dp["tot_transport"],dp["tot_toll"],dp["grand_total"],dp["final_park"]]
        fmts=[None,None,NF,None,None,None,"0.0",MF,None,None,None,NF,"0.000",MF,MF,MF,MF,None]
        for ci,(v,f) in enumerate(zip(vals,fmts),1): _cell(ws1,ri,ci,v,fmt=f,fh=alt)

    tr = len(delivery_plans)+3
    ws1.cell(row=tr,column=1,value="TOTALS").font=_font(bold=True)
    for col,fmt in [(12,NF),(13,"0.000"),(14,MF),(15,MF),(16,MF),(17,MF)]:
        c=ws1.cell(row=tr,column=col)
        c.value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{tr-1})"
        c.number_format=fmt; c.font=_font(bold=True)
        c.border=_border(); c.fill=_fill(C_BLUE_LIGHT); c.alignment=_center()
    _widths(ws1,[9,9,13,34,10,26,14,16,30,30,30,15,14,18,18,13,20,34])

    # ── Sheet 2 : Step-by-step ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Delivery Sequence Detail"); ws2.freeze_panes="A3"
    _title(ws2,"A1:M1",f"DELIVERY SEQUENCE – STEP-BY-STEP   ({today_str})",C_GREEN_DARK)
    h2=["Truck\nID","Truck\nType","Step\n#","Location\nType","Location / Station Name",
        "Qty Delivered\n(Lt)","Qty Delivered\n(MT)","Dist from\nPrev (km)",
        "Toll\nThis Leg (₹)","Transport\nCost (₹)","Purchase\nCost (₹)",
        "Leg Cost\n(₹)","Cumulative\nCost (₹)"]
    for ci,h in enumerate(h2,1): _hdr(ws2,2,ci,h,fh=C_GREEN_DARK)
    ws2.row_dimensions[2].height=44

    row2=3
    for dp in delivery_plans:
        cum=0.0
        rows=[
            {"lt":"🚚 INITIAL PARK","nm":dp["initial_park"],"ql":None,"qm":None,
             "dt":None,"tl":None,"tc":None,"pc":None,"leg":None,"cum":None,"fh":C_BLUE_LIGHT},
            {"lt":"⛽ SOURCE","nm":f"{dp['source_name']}  ({dp['source_id']})",
             "ql":None,"qm":None,"dt":dp["tk_src_dist"],"tl":dp["tk_src_toll"],
             "tc":None,"pc":None,"leg":dp["tk_src_toll"],"cum":None,"fh":C_YELLOW},
        ]
        for i,s in enumerate(dp["stops"]):
            leg=s["transport_cost"]+s["purchase_cost"]+s["toll"]; cum+=leg
            rows.append({"lt":f"📍 STOP {i+1}","nm":s["station"],
                         "ql":s["needed_lt"],"qm":s["needed_mt"],
                         "dt":s["dist_km"],"tl":s["toll"],
                         "tc":s["transport_cost"],"pc":s["purchase_cost"],
                         "leg":round(leg,2),"cum":round(cum,2),"fh":C_GREEN_LIGHT})
        rows.append({"lt":"🏁 FINAL PARK","nm":dp["final_park"],
                     "ql":None,"qm":None,"dt":None,"tl":None,
                     "tc":None,"pc":None,"leg":None,"cum":None,"fh":C_ORANGE_LIGHT})
        for si,sr in enumerate(rows,1):
            vs=[dp["truck_id"],dp["truck_type"],si,sr["lt"],sr["nm"],
                sr["ql"],sr["qm"],sr["dt"],sr["tl"],sr["tc"],sr["pc"],sr["leg"],sr["cum"]]
            fs=[None,None,None,None,None,NF,"0.000","0.1",MF,MF,MF,MF,MF]
            for ci,(v,f) in enumerate(zip(vs,fs),1): _cell(ws2,row2,ci,v,fmt=f,fh=sr["fh"])
            row2+=1
        row2+=1
    _widths(ws2,[9,9,7,20,42,16,14,14,14,16,18,18,20])

    # ── Sheet 3 : Fleet End-of-Day Status ─────────────────────────────────────
    ws3=wb.create_sheet("Fleet End-of-Day Status"); ws3.freeze_panes="A3"
    _title(ws3,"A1:F1",f"FLEET END-OF-DAY STATUS – ALL 30 TRUCKS   ({today_str})",C_PURPLE_DARK)
    h3=["Truck\nID","Truck\nType","Status",
        "Initial Parked At\n(Start of This Run)",
        "Final Parked At\n(End of This Run)","Notes"]
    for ci,h in enumerate(h3,1): _hdr(ws3,2,ci,h,fh=C_PURPLE_DARK)
    ws3.row_dimensions[2].height=44

    for ri,ts in enumerate(fleet_status,3):
        alt=C_GREY_LIGHT if ri%2==0 else C_WHITE
        sf =C_DEPLOYED if ts["status"]=="DEPLOYED" else C_STANDBY
        note="✅ Delivered today" if ts["status"]=="DEPLOYED" else "– On standby"
        for ci,v in enumerate([ts["truck_id"],ts["type"],ts["status"],
                                 ts["initial_park"],ts["final_park"],note],1):
            _cell(ws3,ri,ci,v,fh=(sf if ci==3 else alt))

    last3=len(fleet_status)+4
    dn=sum(1 for t in fleet_status if t["status"]=="DEPLOYED")
    sn=len(fleet_status)-dn
    for label,val,row in [("DEPLOYED",dn,last3),("STANDBY",sn,last3+1)]:
        ws3.cell(row=row,column=1,value=label).font=_font(bold=True)
        ws3.cell(row=row,column=2,value=val).font=_font(bold=True)
    _widths(ws3,[9,10,12,44,44,32])

    # ── Sheet 4 : Cost Summary ─────────────────────────────────────────────────
    ws4=wb.create_sheet("Cost Summary"); ws4.freeze_panes="A3"
    _title(ws4,"A1:G1",f"COST SUMMARY – TODAY'S DISPATCH   ({today_str})",C_ORANGE_DARK)
    h4=["Truck\nID","Source\nID","Stations Served (Delivery Sequence)",
        "Purchase\nCost (₹)","Transport\nCost (₹)","Toll\nCost (₹)","Grand Total\nCost (₹)"]
    for ci,h in enumerate(h4,1): _hdr(ws4,2,ci,h,fh=C_ORANGE_DARK)
    ws4.row_dimensions[2].height=44

    for ri,dp in enumerate(delivery_plans,3):
        alt=C_ORANGE_LIGHT if ri%2==0 else C_WHITE
        seq=" → ".join(s["station"] for s in dp["stops"])
        for ci,(v,f) in enumerate(zip([dp["truck_id"],dp["source_id"],seq,
            dp["tot_purchase"],dp["tot_transport"],dp["tot_toll"],dp["grand_total"]],
            [None,None,None,MF,MF,MF,MF]),1):
            _cell(ws4,ri,ci,v,fmt=f,fh=alt)

    tr4=len(delivery_plans)+3
    ws4.cell(row=tr4,column=1,value="TOTALS").font=_font(bold=True)
    for col,fmt in [(4,MF),(5,MF),(6,MF),(7,MF)]:
        c=ws4.cell(row=tr4,column=col)
        c.value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{tr4-1})"
        c.number_format=fmt; c.font=_font(bold=True)
        c.border=_border(); c.fill=_fill(C_ORANGE_LIGHT); c.alignment=_center()
    _widths(ws4,[9,10,65,20,20,16,22])

    wb.save(output_path)
    print(f"\n[6/6] Excel saved  →  {output_path}")
    print("=" * 65)

# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    delivery_plans, fleet_status, trucks = run_optimization()
    if delivery_plans is None:
        sys.exit(0)

    date_str    = datetime.today().strftime("%Y%m%d_%H%M")
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, f"LPG_DISPATCH_PLAN_{date_str}.xlsx")
    outputs_dir = "/mnt/user-data/outputs"
    if os.path.isdir(outputs_dir):
        output_path = os.path.join(outputs_dir, f"LPG_DISPATCH_PLAN_{date_str}.xlsx")

    build_excel(delivery_plans, fleet_status, trucks, output_path)

    print("\n📋 QUICK SUMMARY")
    print(f"   Runs planned : {len(delivery_plans)}")
    print(f"   Trucks used  : {sum(1 for t in fleet_status if t['status']=='DEPLOYED')} / {len(trucks)}")
    print(f"   Total LPG    : {sum(dp['total_lt'] for dp in delivery_plans):,.0f} Litres")
    print(f"   Total cost   : ₹{sum(dp['grand_total'] for dp in delivery_plans):,.2f}")
    print(f"\n   Output       : {output_path}")
    print(f"   Positions    : {positions_path()}")
    print()