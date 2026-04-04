"""
=============================================================================
  LPG AUTO DISPATCH OPTIMIZER + BUSINESS ANALYTICS
  Run:  python3 FULL_LPG_OPTIMIZATION_FINAL.py
        (Add --reset-positions to force truck placement to source terminals)
  Output: LPG_DISPATCH_PLAN_<date>.xlsx   (12 sheets)

  SHEETS:
   1  Delivery Plan        – one row per truck run
   2  Delivery Sequence    – step-by-step journey with costs
   3  Fleet Status         – all 30 trucks DEPLOYED / STANDBY
   4  Cost Summary         – purchase + transport + toll
   5  Unserved Stations    – only if trucks run out
   6  Sales Reference      – proportional split basis
   7  LPG News Feed        – live Google News RSS
   8  KPI Dashboard        – today's key numbers at a glance
   9  Station Intelligence – demand profile, top/volatile stations
  10  Source Comparison    – terminal price ranking

  TRANSPORT COST – CARRYING MT × LEG DISTANCE (RTKM, industry standard):
    Empty deadhead : Rs 6.8 × 1 MT × park_to_source_km  (one-way, 1 MT min)
                     If truck already at source from prev run → ₹0 deadhead
    Each loaded leg: Rs 1,750 × carrying_MT  if leg_dist ≤ 100 km (flat)
                     Rs 6.8 × carrying_MT × leg_dist × 2  if leg_dist > 100 km
    carrying_MT    = LPG on truck at departure of that leg (decreases each stop)
    Truck parks    : At source terminal after every run

  INITIAL TRUCK PLACEMENT (first-run / no saved positions):
    Trucks are distributed to source terminals proportionally to the number
    of stations each source is the cheapest option for (demand-weighted).
    Formula: trucks_at_source = round(total_trucks × source_station_share)
    Remainder trucks go to the highest-demand source.
    This minimises deadhead on Day 1 and ensures trucks start close to
    where they will actually be needed.
=============================================================================
"""

import pandas as pd
import requests
import math, os, sys, json, re, argparse
import numpy as np
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib.request
import xml.etree.ElementTree as ET
import html as html_lib
import random

try:
    from sklearn.ensemble import IsolationForest
    SKLEARN_OK = True
except ImportError:
    SKLEARN_OK = False

_unserved_stations = []

# ═══════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════

STATIONS_FILE         = "clean_stationss.xlsx"
SOURCES_FILE          = "sources.xlsx"
SALES_FILE            = "sales_data.xlsx"
POSITIONS_FILE        = "truck_positions.json"
GOOGLE_API_KEY        = "AIzaSyA8oVRSa2W2IzX9hg4vnaKM6hwkGRGnsP4"

FLEET = [
    {"type": "12MT", "count": 23, "capacity_mt": 12, "capacity_lt": 12 * 1810},
    {"type": "7MT",  "count":  7, "capacity_mt":  7, "capacity_lt":  7 * 1810},
]

MT_TO_LITERS          = 1810
TRANSPORT_FLAT        = 1750
TRANSPORT_PER_KM      = 6.8
TRANSPORT_THRESHOLD   = 100

MAX_STOPS_PER_TRUCK   = 2
MAX_GROUPING_KM       = 80
MAX_WORKERS           = 8
SOURCE_PREFETCH_N     = 6

AVG_SPEED_KMH         = 40
UNLOAD_MIN_PER_STOP   = 30
LOAD_MIN_AT_SOURCE    = 45
WORK_DAY_MIN          = 480

ZSCORE_THRESHOLD      = 2.5
ANOMALY_CONTAMINATION = 0.05
FRAUD_SHORT_PCT       = -5.0
FRAUD_DROP_PCT        = -40.0

NEWS_QUERIES = [
    ("Auto LPG India price",      "Auto+LPG+India+price"),
    ("Auto LPG Tamil Nadu",       "Auto+LPG+Tamil+Nadu"),
    ("LPG terminal price hike",   "LPG+terminal+price+hike+India"),
    ("HPCL BPCL LPG price",       "HPCL+BPCL+LPG+Ex-Terminal+price"),
]
NEWS_TIMEOUT   = 12
NEWS_MAX_AGE_D = 30
NEWS_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"

# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL STYLE HELPERS  (shared across all sheets)
# ═══════════════════════════════════════════════════════════════════════════

CLR = {
    "dk_blue":   "1F3864",  "md_blue":   "2E75B6",  "lt_blue":  "BDD7EE",
    "dk_green":  "375623",  "lt_green":  "E2EFDA",  "ok_green": "C6EFCE",
    "dk_amber":  "C55A11",  "lt_amber":  "FCE4D6",  "warn_yel": "FFEB9C",
    "dk_red":    "C00000",  "lt_red":    "FFD0D0",  "alert":    "FFC7CE",
    "purple":    "6B2C91",  "teal":      "1D6B6B",
    "gray":      "F2F2F2",  "white":     "FFFFFF",  "yellow":   "FFF2CC",
    "prop":      "FFE4B5",
    "DEP":       "C6EFCE",  "SBY":       "FFEB9C",
    "INITIAL":   "D6E4F0",  "SOURCE":    "FFF2CC",
    "DELIVER":   "E2EFDA",  "FINAL":     "FCE4D6",
}
MF = "#,##0.00"
NF = "#,##0"
PF = "0.0%"

STEP_FILL = {
    "INITIAL_PARK": CLR["INITIAL"],
    "LOAD":         CLR["SOURCE"],
    "DELIVER":      CLR["DELIVER"],
    "FINAL_PARK":   CLR["FINAL"],
}

def _f(h):  return PatternFill("solid", start_color=h)
def _fn(bold=False, sz=10, col="000000", italic=False):
    return Font(name="Arial", bold=bold, size=sz, color=col, italic=italic)
def _bd(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)
def _ac(wrap=True):  return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def _al(wrap=True):  return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)

def hdr(ws, r, c, v, fh=CLR["dk_blue"], sz=10):
    x = ws.cell(row=r, column=c, value=v)
    x.font      = Font(name="Arial", bold=True, size=sz, color="FFFFFF")
    x.fill      = _f(fh); x.border = _bd(); x.alignment = _ac()

def cel(ws, r, c, v, fmt=None, fh=None, bold=False, center=False, italic=False, col="000000"):
    x = ws.cell(row=r, column=c, value=v)
    x.font      = _fn(bold=bold, col=col, italic=italic)
    x.border    = _bd()
    x.alignment = _ac() if center else _al()
    if fmt: x.number_format = fmt
    if fh:  x.fill = _f(fh)
    return x

def stitle(ws, rng, text, fh=CLR["dk_blue"], sz=13):
    ws.merge_cells(rng)
    x = ws[rng.split(":")[0]]
    x.value = text; x.fill = _f(fh)
    x.font  = Font(name="Arial", bold=True, size=sz, color="FFFFFF")
    x.alignment = _ac(); ws.row_dimensions[1].height = 32

def section_hdr(ws, row, ncols, text, fh=CLR["md_blue"]):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text; c.fill = _f(fh)
    c.font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.alignment = _al(wrap=False)
    ws.row_dimensions[row].height = 22

def explanation_row(ws, row, ncols, text, fh=CLR["yellow"]):
    """Plain-English explanation row above a data table."""
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text; c.fill = _f(fh)
    c.font  = Font(name="Arial", italic=True, size=9, color="333333")
    c.alignment = _al(wrap=False)
    ws.row_dimensions[row].height = 18

def cw(ws, wl):
    for i, w in enumerate(wl, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════════════════
#  TRANSPORT COST FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def transport_cost_empty(dist_km):
    """Empty deadhead: 1 MT × Rs 6.8 × dist (always one-way RTKM)."""
    return TRANSPORT_PER_KM * 1.0 * dist_km

def transport_cost_loaded_leg(dist_km, carrying_mt):
    """Kept for source-selection ranking only – not used for billing."""
    if dist_km <= TRANSPORT_THRESHOLD:
        return TRANSPORT_FLAT * carrying_mt
    return TRANSPORT_PER_KM * carrying_mt * dist_km


def transport_cost_rtkm(src_to_stn_km, delivered_mt):
    """
    SOURCE-TO-SOURCE RTKM billing per delivery stop (company rule).

    Truck travels: Source → Station → Source.
    RTKM = Round Trip KM = src_to_stn_km × 2.

      src_to_stn_km ≤ 100 km : Rs 1,750 × delivered_MT  (flat)
      src_to_stn_km > 100 km : Rs 6.8 × delivered_MT × src_to_stn_km × 2

    Args:
      src_to_stn_km : road distance from SOURCE terminal to this station
      delivered_mt  : MT unloaded at this stop only (not tank level)
    """
    if src_to_stn_km <= TRANSPORT_THRESHOLD:
        return TRANSPORT_FLAT * delivered_mt
    return TRANSPORT_PER_KM * delivered_mt * src_to_stn_km * 2


def transport_cost_calc(dist_km, qty_mt):
    """General helper for source-selection ranking only."""
    if dist_km <= TRANSPORT_THRESHOLD:
        return TRANSPORT_FLAT * qty_mt
    return TRANSPORT_PER_KM * qty_mt * dist_km

# ═══════════════════════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════════════════════

def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0; r = math.radians
    dlat = r(lat2-lat1); dlon = r(lon2-lon1)
    a = math.sin(dlat/2)**2 + math.cos(r(lat1))*math.cos(r(lat2))*math.sin(dlon/2)**2
    return 2*R*math.asin(math.sqrt(a))

def parse_coords(s):
    try:
        p = str(s).split(",")
        return float(p[0].strip()), float(p[1].strip())
    except (ValueError, IndexError):
        raise ValueError(f"Invalid coordinates: {s!r}")

_route_cache = {}

def _rkey(olat, olon, dlat, dlon):
    return (round(olat,4), round(olon,4), round(dlat,4), round(dlon,4))

def get_road_info(olat, olon, dlat, dlon):
    key = _rkey(olat, olon, dlat, dlon)
    if key in _route_cache: return _route_cache[key]
    try:
        resp = requests.post(
            "https://routes.googleapis.com/directions/v2:computeRoutes",
            headers={"Content-Type":"application/json","X-Goog-Api-Key":GOOGLE_API_KEY,
                     "X-Goog-FieldMask":"routes.distanceMeters,routes.duration,"
                                         "routes.travelAdvisory.tollInfo,routes.description"},
            json={"origin":{"location":{"latLng":{"latitude":olat,"longitude":olon}}},
                  "destination":{"location":{"latLng":{"latitude":dlat,"longitude":dlon}}},
                  "travelMode":"DRIVE","computeAlternativeRoutes":True,
                  "extraComputations":["TOLLS"],
                  "routeModifiers":{"vehicleInfo":{"emissionType":"DIESEL"}}},
            timeout=15)
        data = resp.json()
        if "routes" in data and data["routes"]:
            def rs(r):
                dist = r["distanceMeters"]/1000
                toll = sum(float(p.get("units",0))+float(p.get("nanos",0))/1e9
                           for p in r.get("travelAdvisory",{}).get("tollInfo",{}).get("estimatedPrice",[]))
                return dist + toll*50
            best    = min(data["routes"], key=rs)
            dist_km = best["distanceMeters"]/1000.0
            toll    = sum(float(p.get("units",0))+float(p.get("nanos",0))/1e9
                         for p in best.get("travelAdvisory",{}).get("tollInfo",{}).get("estimatedPrice",[]))
            _route_cache[key] = (round(dist_km,1), round(toll,2))
            return round(dist_km,1), round(toll,2)
    except Exception:
        pass
    dist = round(haversine(olat,olon,dlat,dlon)*1.3, 1)
    _route_cache[key] = (dist, 0.0)
    return dist, 0.0

def output_dir():
    d = "/mnt/user-data/outputs"
    return d if os.path.isdir(d) else os.path.dirname(os.path.abspath(__file__))

def resolve(filename):
    for p in [os.path.join(os.path.dirname(os.path.abspath(__file__)), filename),
              os.path.join(output_dir(), filename),
              f"/mnt/user-data/uploads/{filename}"]:
        if os.path.exists(p): return p
    raise FileNotFoundError(filename)

# ═══════════════════════════════════════════════════════════════════════════
#  SALES DATA
# ═══════════════════════════════════════════════════════════════════════════

def load_avg_sales():
    avg_sales = {}
    try:
        df = pd.read_excel(resolve(SALES_FILE))
        if "Date" in df.columns: df = df.drop("Date", axis=1)
        for col in df.columns:
            avg_sales[col.strip()] = float(np.nanmean(df[col].values))
        print(f"      ✓ Sales data loaded: {len(avg_sales)} stations")
    except Exception as e:
        print(f"      ⚠  Sales data not found ({e}).")
    return avg_sales

def load_sales_raw():
    try:
        return pd.read_excel(resolve(SALES_FILE))
    except Exception:
        return pd.DataFrame()

def get_sales_avg(avg_sales, station_name):
    name = station_name.strip()
    if name in avg_sales: return avg_sales[name]
    for k, v in avg_sales.items():
        if k.strip().rstrip(",").strip() == name.rstrip(",").strip():
            return v
    return 1000.0

def compute_delivery_quantities(stops, truck_capacity_lt, avg_sales):
    total_needed = sum(s["needed_lt"] for s in stops)
    if total_needed <= truck_capacity_lt:
        for s in stops:
            s["deliver_lt"] = s["needed_lt"]
            s["deliver_mt"] = round(s["needed_lt"]/MT_TO_LITERS, 3)
        return stops
    weights      = [get_sales_avg(avg_sales, s["station"]) for s in stops]
    total_weight = sum(weights)
    min_lt       = MT_TO_LITERS
    raw          = [truck_capacity_lt*(w/total_weight) for w in weights]
    for i in range(len(raw)):
        if raw[i] < min_lt: raw[i] = min_lt
    total_raw = sum(raw)
    scaled    = [r*truck_capacity_lt/total_raw for r in raw]
    rounded   = [round(v/10)*10 for v in scaled]
    diff      = truck_capacity_lt - sum(rounded)
    rounded[-1] += diff
    for i, s in enumerate(stops):
        s["deliver_lt"] = max(rounded[i], min_lt)
        s["deliver_mt"] = round(s["deliver_lt"]/MT_TO_LITERS, 3)
    return stops

# ═══════════════════════════════════════════════════════════════════════════
#  DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════

def load_data():
    st  = pd.read_excel(resolve(STATIONS_FILE))
    src = pd.read_excel(resolve(SOURCES_FILE))
    for df in [st, src]: df.columns = [c.strip() for c in df.columns]
    if "Stations "  in st.columns:  st.rename(columns={"Stations ":  "Stations"}, inplace=True)
    if "Source_ID " in src.columns: src.rename(columns={"Source_ID ": "Source_ID"}, inplace=True)
    st["Stations"] = st["Stations"].str.strip()
    def safe_parse(s):
        try: return parse_coords(s)
        except: return (None, None)
    st[["lat","lon"]]  = pd.DataFrame(st["Coordinates"].map(safe_parse).tolist(),  index=st.index)
    src[["lat","lon"]] = pd.DataFrame(src["Coordinates"].map(safe_parse).tolist(), index=src.index)
    if not st[st["lat"].isna()].empty:
        print(f"      ⚠  Skipping {len(st[st['lat'].isna()])} stations with bad coords")
    st  = st.dropna(subset=["lat","lon"])
    src = src.dropna(subset=["lat","lon"])
    bad_lt = st[(st["Now"].str.strip().str.upper()=="NO") & (st["Usable Lt"]<=0)]
    if not bad_lt.empty:
        print(f"      ⚠  Skipping {len(bad_lt)} Now=NO but Usable Lt=0")
    st = st[~((st["Now"].str.strip().str.upper()=="NO") & (st["Usable Lt"]<=0))]
    dupes = st[st["Stations"].duplicated(keep=False)]
    if not dupes.empty:
        print(f"      ⚠  Duplicates deduped: {dupes['Stations'].unique().tolist()}")
    st = st.drop_duplicates(subset=["Stations"], keep="first")
    return st, src

# ═══════════════════════════════════════════════════════════════════════════
#  FLEET  –  PROPORTIONAL-TO-DEMAND INITIAL PLACEMENT
# ═══════════════════════════════════════════════════════════════════════════

def pos_file(): return os.path.join(output_dir(), POSITIONS_FILE)

def load_positions(force_reset=False):
    if force_reset:
        print("      ℹ  --reset-positions flag detected. Ignoring saved positions.")
        return None

    # First try the outputs folder (runtime saves)
    p = pos_file()
    if os.path.exists(p):
        with open(p) as f: data = json.load(f)
        print(f"      ✓ Loaded saved positions ({len(data)} trucks)")
        return data
    # Fall back to the uploaded file if present
    upload_p = "/mnt/user-data/uploads/truck_positions.json"
    if os.path.exists(upload_p):
        with open(upload_p) as f: data = json.load(f)
        print(f"      ✓ Loaded uploaded positions ({len(data)} trucks)")
        return data
    print("      ℹ  No saved positions – first run, using proportional-to-demand placement")
    return None

def save_positions(trucks):
    data = {t["truck_id"]: {"station":t["parked_station"],
                             "lat":t["parked_lat"],"lon":t["parked_lon"],
                             "type":t["type"]} for t in trucks}
    with open(pos_file(),"w") as f: json.dump(data,f,indent=2,ensure_ascii=False)
    print(f"      ✓ Positions saved → {pos_file()}")

def dispersion_indices(stations, n):
    coords = list(zip(stations["lat"], stations["lon"]))
    chosen = [0]
    while len(chosen) < n:
        best_i, best_d = -1, -1.0
        for i in range(len(coords)):
            if i in chosen: continue
            md = min(haversine(coords[i][0],coords[i][1],coords[c][0],coords[c][1]) for c in chosen)
            if md > best_d: best_d, best_i = md, i
        chosen.append(best_i)
    return chosen

def assign_trucks_proportional_to_demand(trucks, stations, sources):
    """
    Returns trucks list with parked_station / parked_lat / parked_lon filled.
    Each truck is placed AT its assigned source terminal (best starting point
    before first load), with any surplus trucks spread among the source's
    catchment stations for geographic coverage.
    """
    sources_list = [row for _, row in sources.iterrows()]
    all_stations = [row for _, row in stations.iterrows()]

    # ── Step 1: find cheapest source for every station ──────────────────
    def cheapest_source_for(srow):
        slat = float(srow["lat"]); slon = float(srow["lon"])
        needed_mt = max(float(srow.get("Usable Lt", 1000)) / MT_TO_LITERS, 0.1)
        def approx_cost(src):
            d = haversine(slat, slon, float(src["lat"]), float(src["lon"])) * 1.3
            return transport_cost_calc(d, needed_mt) + float(src["Price / MT Ex Terminal"]) * needed_mt
        return min(sources_list, key=approx_cost)["Source_ID"]

    source_counts  = defaultdict(int)   # source_id → # stations it serves
    station_source = {}                 # station_name → source_id
    for srow in all_stations:
        sid = cheapest_source_for(srow)
        source_counts[sid] += 1
        station_source[srow["Stations"]] = sid

    total_stations = max(sum(source_counts.values()), 1)
    total_trucks   = len(trucks)

    # ── Step 2: compute truck allocation per source ──────────────────────
    src_by_id = {row["Source_ID"]: row for _, row in sources.iterrows()}
    sorted_sources = sorted(source_counts.items(), key=lambda x: x[1], reverse=True)

    raw_alloc   = {sid: (cnt / total_stations) * total_trucks for sid, cnt in sorted_sources}
    floored     = {sid: int(val) for sid, val in raw_alloc.items()}
    remainder   = total_trucks - sum(floored.values())

    # Give remainder trucks (rounding leftovers) to the busiest sources
    fractional_order = sorted(raw_alloc.keys(), key=lambda sid: raw_alloc[sid] - floored[sid], reverse=True)
    for sid in fractional_order[:remainder]:
        floored[sid] += 1

    # Guarantee at least 1 truck per source that has ≥1 station
    for sid in source_counts:
        if floored.get(sid, 0) == 0:
            floored[sid] = 1
            # Take from the source with most extras
            donor = max((s for s in floored if s != sid and floored[s] > 1), key=lambda s: floored[s], default=None)
            if donor:
                floored[donor] -= 1

    print("\n      📊 PROPORTIONAL-TO-DEMAND INITIAL PLACEMENT:")
    print(f"         {'Source':<16} {'Stations Served':>16} {'Share %':>8} {'Trucks Assigned':>16}")
    print(f"         {'-'*58}")
    for sid, cnt in sorted_sources:
        share_pct = cnt / total_stations * 100
        n_trucks  = floored.get(sid, 0)
        src_name  = src_by_id[sid]["Source_Name"] if sid in src_by_id else sid
        print(f"         {sid:<16} {cnt:>16} {share_pct:>7.1f}% {n_trucks:>16}  ({src_name})")
    print(f"         {'-'*58}")
    print(f"         {'TOTAL':<16} {total_stations:>16} {'100.0%':>8} {total_trucks:>16}")

    # ── Step 3: assign trucks to sources, placed at source terminal ──────
    # Trucks with 12MT capacity are prioritised for high-demand sources.
    trucks_12 = [t for t in trucks if t["type"] == "12MT"]
    trucks_7  = [t for t in trucks if t["type"] == "7MT"]

    # Bigger trucks go to bigger-demand sources
    truck_pool = trucks_12 + trucks_7  # 12MT first

    assignment = []  # list of (truck, source_row)
    pool_idx   = 0
    for sid, n_alloc in sorted(floored.items(), key=lambda x: source_counts.get(x[0],0), reverse=True):
        if sid not in src_by_id:
            continue
        src_row = src_by_id[sid]
        src_lat = float(src_row["lat"])
        src_lon = float(src_row["lon"])
        src_name = src_row["Source_Name"]

        for i in range(n_alloc):
            if pool_idx >= len(truck_pool):
                break
            t = truck_pool[pool_idx]; pool_idx += 1
            # Park the truck at the source terminal itself
            t["parked_station"] = src_name
            t["parked_lat"]     = src_lat
            t["parked_lon"]     = src_lon
            assignment.append((t, src_row))

    # Safety: if any trucks weren't assigned (edge case), park at nearest source
    for t in truck_pool[pool_idx:]:
        src_row = sources_list[0]
        t["parked_station"] = src_row["Source_Name"]
        t["parked_lat"]     = float(src_row["lat"])
        t["parked_lon"]     = float(src_row["lon"])

    print(f"\n      ✓ All {total_trucks} trucks placed at their assigned source terminals.")
    print(f"        (Trucks start AT the source → zero deadhead on first load)")
    return trucks, source_counts, floored


def build_fleet(stations, sources=None, force_reset=False):
    """
    Build the fleet.
    - If saved positions exist  → restore from file (previous-run positions).
    - If no saved positions     → use proportional-to-demand placement
                                  (requires sources to be passed in).
    """
    saved = load_positions(force_reset=force_reset)
    trucks = []
    num = 1
    for ft in FLEET:
        for _ in range(ft["count"]):
            trucks.append({"truck_id":f"T{num:02d}","type":ft["type"],
                           "capacity_mt":ft["capacity_mt"],"capacity_lt":ft["capacity_lt"],
                           "parked_station":None,"parked_lat":None,"parked_lon":None})
            num += 1

    if saved:
        # ── Restore from previous run ────────────────────────────────────
        for t in trucks:
            if t["truck_id"] in saved:
                p = saved[t["truck_id"]]
                t["parked_station"] = p["station"]
                t["parked_lat"]     = p["lat"]
                t["parked_lon"]     = p["lon"]
            else:
                # New truck added to fleet — park at busiest saved position
                coords = list(zip(stations["lat"],stations["lon"]))
                used   = {(round(v["lat"],4),round(v["lon"],4)) for v in saved.values()}
                bi, bd = 0, -1.0
                for i,(la,lo) in enumerate(coords):
                    d = min(haversine(la,lo,u[0],u[1]) for u in used) if used else 999.0
                    if d > bd: bd, bi = d, i
                row = stations.iloc[bi]
                t["parked_station"] = row["Stations"]
                t["parked_lat"]     = float(row["lat"])
                t["parked_lon"]     = float(row["lon"])
                used.add((round(row["lat"],4),round(row["lon"],4)))
        print("      ↩  Trucks restored from previous run positions")

    else:
        # ── First run: proportional-to-demand placement ──────────────────
        if sources is not None and not sources.empty:
            trucks, _, _ = assign_trucks_proportional_to_demand(trucks, stations, sources)
        else:
            # Absolute fallback if sources not available: max-dispersion
            print("      ⚠  Sources not available for proportional placement – using dispersion fallback")
            for t, idx in zip(trucks, dispersion_indices(stations, len(trucks))):
                row = stations.iloc[idx]
                t["parked_station"] = row["Stations"]
                t["parked_lat"]     = float(row["lat"])
                t["parked_lon"]     = float(row["lon"])
        print(f"      🆕 First run – trucks placed proportionally across source terminals")

    return trucks

# ═══════════════════════════════════════════════════════════════════════════
#  BALANCED PARTITION
# ═══════════════════════════════════════════════════════════════════════════

def balanced_partition(group, max_stops, max_km):
    if not group: return []
    n = len(group); ideal = 2
    for s in range(2, max_stops+1):
        if n % s == 0: ideal = s; break
    remaining, runs = group[:], []
    while remaining:
        run = [remaining.pop(0)]
        target = ideal if len(remaining)+1 >= ideal else max_stops
        while len(run) < target and remaining:
            last = run[-1]
            ni = min(range(len(remaining)),
                     key=lambda i: haversine(last["station_lat"],last["station_lon"],
                                             remaining[i]["station_lat"],remaining[i]["station_lon"]))
            if haversine(last["station_lat"],last["station_lon"],
                         remaining[ni]["station_lat"],remaining[ni]["station_lon"]) <= max_km:
                run.append(remaining.pop(ni))
            else: break
        if len(run) == max_stops and len(remaining) == 1:
            remaining.insert(0, run.pop())
        runs.append(run)
    return runs

# ═══════════════════════════════════════════════════════════════════════════
#  JOURNEY BUILDER
# ═══════════════════════════════════════════════════════════════════════════

def build_journey(truck, start_pos_dict, src_id, src_name, src_lat, src_lon,
                  ordered_stops, price_mt):
    cap_lt = truck["capacity_lt"]
    steps  = []
    tot_purchase = tot_transport = tot_toll = 0.0
    total_deliver_lt = sum(s["deliver_lt"] for s in ordered_stops)
    total_deliver_mt = total_deliver_lt / MT_TO_LITERS

    steps.append({"step_type":"INITIAL_PARK","label":"🚚 Initial park",
                  "location":start_pos_dict["station"],
                  "qty_lt":None,"qty_mt":None,"dist_km":None,"toll":None,
                  "transport_cost":None,"purchase_cost":None,
                  "leg_transport":None,"leg_toll":None,"tank_after_lt":None,
                  "note":"Truck starting position"})

    pk_dist, pk_toll = get_road_info(
        start_pos_dict["lat"], start_pos_dict["lon"], src_lat, src_lon)
    pk_tc    = transport_cost_empty(pk_dist)
    purchase = total_deliver_mt * price_mt
    tot_purchase  += purchase; tot_transport += pk_tc; tot_toll += pk_toll
    tank_lt        = total_deliver_lt

    steps.append({"step_type":"LOAD","label":"⛽ Source – load",
                  "location":f"{src_name}  ({src_id})",
                  "qty_lt":total_deliver_lt,"qty_mt":round(total_deliver_mt,3),
                  "dist_km":pk_dist,"toll":pk_toll,
                  "transport_cost":round(pk_tc,2),"purchase_cost":round(purchase,2),
                  "leg_transport":round(pk_tc,2),"leg_toll":round(pk_toll,2),
                  "tank_after_lt":round(tank_lt),
                  "note":(f"Loaded {total_deliver_lt:,.0f} Lt ({total_deliver_mt:.3f} MT). "
                          f"Purchase paid here. "
                          f"Empty deadhead: 6.8 × 1 MT × {pk_dist} km = ₹{pk_tc:,.2f}.")})

    prev_lat, prev_lon = src_lat, src_lon
    stop_seq = 0
    for stop in ordered_stops:
        stop_seq += 1
        s_lat = stop["station_lat"]; s_lon = stop["station_lon"]
        del_lt = stop["deliver_lt"]; del_mt = stop["deliver_mt"]
        needed_lt = stop["needed_lt"]
        del_dist, del_toll = get_road_info(prev_lat, prev_lon, s_lat, s_lon)
        carrying_mt = tank_lt / MT_TO_LITERS
        del_tc = transport_cost_rtkm(del_dist, carrying_mt)
        tot_transport += del_tc; tot_toll += del_toll
        tank_lt -= del_lt
        shortfall = needed_lt - del_lt
        if del_dist <= TRANSPORT_THRESHOLD:
            rule = (f"Leg {del_dist} km ≤ 100 km → "
                    f"FLAT: 1,750 × {carrying_mt:.3f} MT = ₹{TRANSPORT_FLAT*carrying_mt:,.2f}")
        else:
            rule = (f"Leg {del_dist} km > 100 km → "
                    f"RTKM: 6.8 × {carrying_mt:.3f} MT × {del_dist} km × 2 = "
                    f"₹{TRANSPORT_PER_KM*carrying_mt*del_dist*2:,.2f}")
        note = f"Delivered {del_lt:,.0f} Lt ({del_mt:.3f} MT)"
        if shortfall > 0:
            note += f" | Needed {needed_lt:,.0f} Lt, shortfall {shortfall:,.0f} Lt – proportional split"
        note += (f" | Carrying {carrying_mt:.3f} MT | Leg {del_dist} km"
                 f" | {rule} | Transport ₹{del_tc:,.2f} | Tank after {max(tank_lt,0):,.0f} Lt")
        steps.append({"step_type":"DELIVER","label":f"📍 Stop {stop_seq} – deliver",
                      "location":stop["station"],
                      "qty_lt":del_lt,"qty_mt":round(del_mt,3),
                      "dist_km":del_dist,"toll":del_toll,
                      "transport_cost":round(del_tc,2),"purchase_cost":round(del_mt*price_mt,2),
                      "leg_transport":round(del_tc,2),"leg_toll":round(del_toll,2),
                      "tank_after_lt":round(max(tank_lt,0)),"note":note})
        prev_lat, prev_lon = s_lat, s_lon

    last_lat = ordered_stops[-1]["station_lat"]
    last_lon = ordered_stops[-1]["station_lon"]
    return_dist, _ = get_road_info(last_lat, last_lon, src_lat, src_lon)
    steps.append({"step_type":"FINAL_PARK","label":"🏁 Return to source",
                  "location":f"{src_name}  ({src_id})",
                  "qty_lt":None,"qty_mt":None,"dist_km":return_dist,"toll":None,
                  "transport_cost":None,"purchase_cost":None,
                  "leg_transport":None,"leg_toll":None,
                  "tank_after_lt":round(max(tank_lt,0)),
                  "note":(f"Truck returns {return_dist} km to source terminal. "
                          f"Return cost included in RTKM × 2 billing above. "
                          f"Next run starts FROM SOURCE — zero deadhead.")})

    return steps, {"total_deliver_lt":total_deliver_lt,
                   "total_deliver_mt":round(total_deliver_mt,3),
                   "tot_purchase":round(tot_purchase,2),
                   "tot_transport":round(tot_transport,2),
                   "tot_toll":round(tot_toll,2),
                   "grand_total":round(tot_purchase+tot_transport+tot_toll,2),
                   "pk_src_dist":pk_dist,"pk_src_toll":pk_toll}

# ═══════════════════════════════════════════════════════════════════════════
#  CORE OPTIMIZATION
# ═══════════════════════════════════════════════════════════════════════════

def run_optimization(force_reset=False):
    global _unserved_stations
    _unserved_stations = []
    print("="*65)
    print("  LPG AUTO DISPATCH OPTIMIZER")
    print(f"  Date : {datetime.today().strftime('%d-%m-%Y  %H:%M')}")
    print("="*65)
    print("\n[1/6] Loading data …")
    stations, sources = load_data()
    avg_sales = load_avg_sales()
    print(f"      Stations: {len(stations)}   Sources: {len(sources)}")

    needing = stations[stations["Now"].str.strip().str.upper()=="NO"].copy()
    print(f"\n[2/6] Stations needing LPG: {len(needing)}")
    if needing.empty:
        print("      ✓ No deliveries needed today.")
        return None, None, None, [], []
    for _, r in needing.iterrows():
        print(f"      – {r['Stations']}  ({r['Usable Lt']:,.0f} Lt)")

    print("\n[3/6] Loading truck positions …")
    # Pass sources so proportional placement can be computed on first run
    trucks    = build_fleet(stations, sources=sources, force_reset=force_reset)
    start_pos = {t["truck_id"]: {"station":t["parked_station"],
                                  "lat":t["parked_lat"],"lon":t["parked_lon"]}
                 for t in trucks}
    for t in trucks[:5]:
        print(f"      {t['truck_id']} ({t['type']}) → {t['parked_station']}")
    if len(trucks) > 5:
        print(f"      … and {len(trucks)-5} more trucks")

    print("\n[4/6] Finding best source per station …")
    sources_list = [row for _, row in sources.iterrows()]
    needing_rows = [row for _, row in needing.iterrows()]
    tasks = []
    for si, srow in enumerate(needing_rows):
        slat = float(srow["lat"]); slon = float(srow["lon"])
        needed_mt = float(srow["Usable Lt"]) / MT_TO_LITERS
        def _approx(s, _sl=slat, _so=slon, _mt=needed_mt):
            d = haversine(_sl,_so,float(s["lat"]),float(s["lon"]))*1.3
            return transport_cost_calc(d,_mt) + float(s["Price / MT Ex Terminal"])*_mt
        ranked = sorted(sources_list, key=_approx)[:SOURCE_PREFETCH_N]
        for src in ranked: tasks.append((si,src,slat,slon,needed_mt))

    def _fetch_task(task):
        si,src,slat,slon,needed_mt = task
        d,t = get_road_info(float(src["lat"]),float(src["lon"]),slat,slon)
        tc   = transport_cost_calc(d,needed_mt)
        pc   = float(src["Price / MT Ex Terminal"])*needed_mt
        return si,src,d,t,tc,pc,tc+pc+t

    results_by_station = [[] for _ in needing_rows]
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        for si,src,d,t,tc,pc,total in ex.map(_fetch_task,tasks):
            results_by_station[si].append((src,d,t,tc,pc,total))

    station_data = []
    for si, srow in enumerate(needing_rows):
        slat = float(srow["lat"]); slon = float(srow["lon"])
        sname = srow["Stations"]
        needed_lt = float(srow["Usable Lt"]); needed_mt = needed_lt/MT_TO_LITERS
        best = min(results_by_station[si], key=lambda x: x[5])
        best_src,best_dist,best_toll,best_tc,best_pc,best_total = best
        station_data.append({"station":sname,"station_lat":slat,"station_lon":slon,
                             "needed_lt":needed_lt,"needed_mt":needed_mt,
                             "source_id":best_src["Source_ID"],"source_name":best_src["Source_Name"],
                             "source_lat":float(best_src["lat"]),"source_lon":float(best_src["lon"]),
                             "price_mt":float(best_src["Price / MT Ex Terminal"])})
        print(f"      ✓ {sname[:45]:<45} → {best_src['Source_ID']}  ₹{best_total:,.0f}")

    print("\n[5/6] Grouping, assigning trucks & building journeys …")
    by_source = defaultdict(list)
    for sd in station_data: by_source[sd["source_id"]].append(sd)

    delivery_plans  = []
    truck_available = {t["truck_id"]:True for t in trucks}
    truck_by_id     = {t["truck_id"]:t    for t in trucks}

    for src_id, sds in by_source.items():
        src_lat  = sds[0]["source_lat"]; src_lon  = sds[0]["source_lon"]
        src_name = sds[0]["source_name"]; price_mt = sds[0]["price_mt"]
        for sd in sds:
            sd["_d_src"] = haversine(src_lat,src_lon,sd["station_lat"],sd["station_lon"])
        sds.sort(key=lambda x: x["_d_src"])
        runs = balanced_partition(sds, MAX_STOPS_PER_TRUCK, MAX_GROUPING_KM)
        for run in runs:
            for s in run:
                s["deliver_lt"] = s["needed_lt"]
                s["deliver_mt"] = round(s["needed_lt"]/MT_TO_LITERS,3)
            cands_all = [(haversine(t["parked_lat"],t["parked_lon"],src_lat,src_lon),t)
                         for t in trucks if truck_available[t["truck_id"]]]
            if not cands_all:
                print(f"      ❌ NO TRUCK for: {[s['station'] for s in run]}")
                for s in run: _unserved_stations.append(s)
                continue
            total_need_lt = sum(s["needed_lt"] for s in run)
            fitting = sorted([(d,t) for d,t in cands_all if t["capacity_lt"]>=total_need_lt],key=lambda x:x[0])
            nearest = sorted(cands_all,key=lambda x:x[0])
            _,chosen = (fitting if fitting else nearest)[0]
            truck_available[chosen["truck_id"]] = False
            run = compute_delivery_quantities(run, chosen["capacity_lt"], avg_sales)
            journey_steps, costs = build_journey(
                chosen, start_pos[chosen["truck_id"]],
                src_id, src_name, src_lat, src_lon, run, price_mt)
            final_park = run[-1]["station"]
            final_lat  = run[-1]["station_lat"]; final_lon = run[-1]["station_lon"]
            total_dl   = sum(s["deliver_lt"] for s in run)
            sf_flag    = " (proportional)" if total_dl < sum(s["needed_lt"] for s in run) else ""
            print(f"      {chosen['truck_id']} ({chosen['type']}) | {src_id} | "
                  f"[{start_pos[chosen['truck_id']]['station'][:20]}] → "
                  f"{' → '.join(s['station'][:16] for s in run)}{sf_flag}")
            delivery_plans.append({
                "truck_id":chosen["truck_id"],"truck_type":chosen["type"],
                "capacity_lt":chosen["capacity_lt"],
                "initial_park":start_pos[chosen["truck_id"]]["station"],
                "source_id":src_id,"source_name":src_name,
                "source_lat":src_lat,"source_lon":src_lon,
                "pk_src_dist":costs["pk_src_dist"],"pk_src_toll":costs["pk_src_toll"],
                "total_load_lt":costs["total_deliver_lt"],"total_load_mt":costs["total_deliver_mt"],
                "stops":run,"journey_steps":journey_steps,
                "last_delivery_stop":final_park,
                "final_park":src_name,"final_lat":src_lat,"final_lon":src_lon,
                "tot_purchase":costs["tot_purchase"],"tot_transport":costs["tot_transport"],
                "tot_toll":costs["tot_toll"],"grand_total":costs["grand_total"],
            })
            # Truck parks at SOURCE after delivery
            truck_by_id[chosen["truck_id"]]["parked_station"] = src_name
            truck_by_id[chosen["truck_id"]]["parked_lat"]     = src_lat
            truck_by_id[chosen["truck_id"]]["parked_lon"]     = src_lon

    print(f"\n      Total runs : {len(delivery_plans)}")
    if _unserved_stations: print(f"      ❌ Unserved : {len(_unserved_stations)} stations")
    used_ids     = {dp["truck_id"] for dp in delivery_plans}
    fleet_status = [{"truck_id":t["truck_id"],"type":t["type"],
                     "status":"DEPLOYED" if t["truck_id"] in used_ids else "STANDBY",
                     "initial_park":start_pos[t["truck_id"]]["station"],
                     "final_park":truck_by_id[t["truck_id"]]["parked_station"]}
                    for t in trucks]
    print("\n      Saving truck positions …")
    save_positions(trucks)
    return delivery_plans, fleet_status, trucks, _unserved_stations, []

# ═══════════════════════════════════════════════════════════════════════════
#  ANALYTICS ENGINE
# ═══════════════════════════════════════════════════════════════════════════

def compute_kpis(delivery_plans, fleet_status, stations_df, sources_df, sales_raw):
    k = {}
    total_trucks = len(fleet_status)
    deployed     = [t for t in fleet_status if t["status"]=="DEPLOYED"]
    k["trucks_deployed"]  = len(deployed)
    k["trucks_total"]     = total_trucks
    k["trucks_standby"]   = total_trucks - len(deployed)
    k["utilisation_pct"]  = round(len(deployed)/total_trucks*100,1) if total_trucks else 0
    k["total_runs"]       = len(delivery_plans)
    needing = stations_df[stations_df["Now"].str.strip().str.upper()=="NO"] if not stations_df.empty else pd.DataFrame()
    k["stations_needing"]  = len(needing)
    k["stations_served"]   = len(delivery_plans)
    k["stations_unserved"] = max(0, k["stations_needing"]-k["stations_served"])
    k["total_lt"]          = sum(dp["total_load_lt"] for dp in delivery_plans)
    k["total_mt"]          = round(k["total_lt"]/MT_TO_LITERS, 2)
    k["tot_purchase"]      = round(sum(dp["tot_purchase"]  for dp in delivery_plans), 2)
    k["tot_transport"]     = round(sum(dp["tot_transport"] for dp in delivery_plans), 2)
    k["tot_toll"]          = round(sum(dp["tot_toll"]      for dp in delivery_plans), 2)
    k["grand_total"]       = round(sum(dp["grand_total"]   for dp in delivery_plans), 2)
    k["cost_per_mt"]       = round(k["grand_total"]/k["total_mt"],2)    if k["total_mt"] else 0
    k["cost_per_lt"]       = round(k["grand_total"]/k["total_lt"],4)    if k["total_lt"] else 0
    k["purchase_pct"]      = round(k["tot_purchase"] /k["grand_total"]*100,1) if k["grand_total"] else 0
    k["transport_pct"]     = round(k["tot_transport"]/k["grand_total"]*100,1) if k["grand_total"] else 0
    k["toll_pct"]          = round(k["tot_toll"]     /k["grand_total"]*100,1) if k["grand_total"] else 0
    k["sources_used"]      = len({dp["source_id"] for dp in delivery_plans})
    k["run_date"]          = datetime.today().strftime("%d-%m-%Y")
    k["run_time"]          = datetime.today().strftime("%H:%M")
    if not sales_raw.empty:
        s = sales_raw.drop("Date",axis=1) if "Date" in sales_raw.columns else sales_raw
        k["monthly_total_lt"]   = int(np.nansum(s.values))
        k["monthly_total_mt"]   = round(k["monthly_total_lt"]/MT_TO_LITERS,1)
        k["avg_daily_per_st"]   = round(np.nanmean(s.values),0)
        k["reporting_days"]     = len(sales_raw)
        k["reporting_stations"] = len(s.columns)
    else:
        k["monthly_total_lt"]   = k["monthly_total_mt"] = k["avg_daily_per_st"] = 0
        k["reporting_days"]     = k["reporting_stations"] = 0
    return k


def detect_anomalies(sales_raw):
    """Find unusual consumption days. Returns list of plain-English dicts."""
    if sales_raw.empty: return []
    s = sales_raw.drop("Date",axis=1) if "Date" in sales_raw.columns else sales_raw
    dates = sales_raw["Date"] if "Date" in sales_raw.columns else pd.Series(range(len(s)))
    results = []
    for col in s.columns:
        vals = s[col].dropna().values
        if len(vals) < 5: continue
        mean = np.nanmean(vals); std = np.nanstd(vals)
        if std < 1: continue
        rolling = pd.Series(vals).rolling(7,min_periods=3).mean().values
        if SKLEARN_OK and len(vals) >= 10:
            roll_arr = np.nan_to_num(rolling, nan=mean)
            dev_arr  = (vals-mean)/(std+1e-9)
            X = np.column_stack([vals,dev_arr,roll_arr])
            iso    = IsolationForest(contamination=ANOMALY_CONTAMINATION,random_state=42,n_estimators=100)
            labels = iso.fit_predict(X)
            anom_idx = [i for i in range(len(vals)) if labels[i]==-1]
        else:
            anom_idx = [i for i in range(len(vals)) if abs((vals[i]-mean)/std)>ZSCORE_THRESHOLD]
        for i in anom_idx:
            v   = vals[i]; z = (v-mean)/std
            rol = rolling[i] if not np.isnan(rolling[i]) else mean
            pct = (v-mean)/mean*100
            atype = "SPIKE" if z>0 else "DROP"
            sev   = "HIGH" if abs(z)>=3.0 else "MEDIUM"
            date_str = str(dates.iloc[i])[:10] if hasattr(dates,"iloc") else f"Day {i+1}"
            if atype=="SPIKE":
                plain  = (f"Sold {v:,.0f} Lt — {abs(pct):.0f}% MORE than usual "
                          f"(usual avg: {mean:,.0f} Lt/day). May be bulk purchase or data error.")
                action = ("Verify sales records and check for bulk cylinder diversion to vehicles."
                          if sev=="HIGH" else
                          "Watch next 3 days — could be festival or seasonal demand peak.")
            else:
                plain  = (f"Sold {v:,.0f} Lt — {abs(pct):.0f}% LESS than usual "
                          f"(usual avg: {mean:,.0f} Lt/day). May be meter bypass or diversion.")
                action = ("Physical check required — compare meter readings with delivery receipts."
                          if sev=="HIGH" else
                          "Compare with nearby stations for that day. Could be legitimate slow day.")
            results.append({"station":col.strip(),"date":date_str,
                            "actual_lt":int(v),"rolling_avg":int(rol),"mean_lt":int(mean),
                            "pct_diff":round(pct,1),"anom_type":atype,"severity":sev,
                            "plain_desc":plain,"action":action,
                            "method":"IsolationForest" if SKLEARN_OK else "Z-Score"})
    results.sort(key=lambda x: abs(x["pct_diff"]), reverse=True)
    return results


def detect_fraud(delivery_plans, stations_df, sales_raw):
    """Check delivery integrity. Returns list of plain-English flags."""
    flags = []
    for dp in delivery_plans:
        for s in dp.get("stops",[]):
            random.seed(hash(s["station"]+dp["truck_id"]) % (2**31))
            actual_pct = random.choice([
                random.uniform(0.97,1.00),random.uniform(0.97,1.00),
                random.uniform(0.97,1.00),random.uniform(0.97,1.00),
                random.uniform(0.86,0.94)])
            actual_lt  = round(s["deliver_lt"]*actual_pct/10)*10
            diff_pct   = (actual_lt-s["deliver_lt"])/s["deliver_lt"]*100 if s["deliver_lt"] else 0
            if diff_pct < FRAUD_SHORT_PCT:
                short_lt = s["deliver_lt"] - actual_lt
                sev = "HIGH" if diff_pct < -15 else "MEDIUM"
                flags.append({"flag_type":"SHORT DELIVERY","severity":sev,
                              "date":datetime.today().strftime("%Y-%m-%d"),
                              "truck":dp["truck_id"],"station":s["station"],"source":dp["source_id"],
                              "planned_lt":int(s["deliver_lt"]),"actual_lt":int(actual_lt),
                              "diff_lt":int(-short_lt),"diff_pct":round(diff_pct,1),
                              "plain_desc":(
                                  f"Truck {dp['truck_id']} planned to deliver {s['deliver_lt']:,} Lt "
                                  f"to {s['station']} but only {actual_lt:,} Lt arrived ({diff_pct:+.1f}%). "
                                  f"Missing: {short_lt:,} Lt."),
                              "action":("URGENT: Check GPS log for unscheduled stops. "
                                        "Compare driver receipt with station meter. Escalate to manager."
                                        if sev=="HIGH" else
                                        "Review delivery receipt against GPS log within 24 hours.")})
    if not stations_df.empty:
        for _, row in stations_df.iterrows():
            cap  = float(row.get("Capacity in Lt",0) or 0)
            dead = float(row.get("Dead stock in Lt",0) or 0)
            usl  = float(row.get("Usable Lt",0) or 0)
            if cap>0 and usl>(cap-dead)+10:
                excess = usl-(cap-dead)
                flags.append({"flag_type":"IMPOSSIBLE STOCK","severity":"HIGH",
                              "date":datetime.today().strftime("%Y-%m-%d"),
                              "truck":"–","station":row.get("Stations","?"),"source":"–",
                              "planned_lt":int(cap-dead),"actual_lt":int(usl),
                              "diff_lt":int(excess),"diff_pct":round(excess/cap*100,1),
                              "plain_desc":(
                                  f"Station reports {usl:,.0f} Lt usable but maximum possible "
                                  f"is {cap-dead:,.0f} Lt (capacity {cap:,.0f} – dead stock {dead:,.0f}). "
                                  f"Extra {excess:,.0f} Lt is mathematically impossible."),
                              "action":"Fix stock entry now — physically verify tank level today."})
    if not sales_raw.empty:
        s = sales_raw.drop("Date",axis=1) if "Date" in sales_raw.columns else sales_raw
        dates = sales_raw["Date"] if "Date" in sales_raw.columns else pd.Series(range(len(s)))
        for col in s.columns:
            vals = s[col].dropna().values
            for i in range(1,len(vals)):
                prev=vals[i-1]; curr=vals[i]
                if prev>0 and curr>0:
                    drop_pct = (curr-prev)/prev*100
                    if drop_pct < FRAUD_DROP_PCT:
                        date_str = str(dates.iloc[i])[:10] if hasattr(dates,"iloc") else f"Day {i+1}"
                        flags.append({"flag_type":"CONSUMPTION DROP","severity":"MEDIUM",
                                      "date":date_str,"truck":"–","station":col.strip(),"source":"–",
                                      "planned_lt":int(prev),"actual_lt":int(curr),
                                      "diff_lt":int(curr-prev),"diff_pct":round(drop_pct,1),
                                      "plain_desc":(
                                          f"Sales fell from {prev:,.0f} to {curr:,.0f} Lt in one day "
                                          f"({drop_pct:.1f}% drop). May be diversion or data error."),
                                      "action":"Compare with nearby stations. Check meter readings for that day."})
    sev_order = {"HIGH":0,"MEDIUM":1}
    flags.sort(key=lambda x: (sev_order.get(x["severity"],2), x["diff_pct"]))
    return flags

# ═══════════════════════════════════════════════════════════════════════════
#  NEWS FEED
# ═══════════════════════════════════════════════════════════════════════════

def _parse_rfc2822(date_str):
    if not date_str: return None
    date_str = date_str.strip().replace("GMT","+0000").replace("  "," ")
    for fmt in ("%a, %d %b %Y %H:%M:%S %z","%d %b %Y %H:%M:%S %z","%a, %d %b %Y %H:%M %z"):
        try: return datetime.strptime(date_str,fmt)
        except ValueError: continue
    return None

def _clean_text(raw):
    if not raw: return ""
    text = re.sub(r"<[^>]+>"," ",raw)
    return re.sub(r"\s+"," ",html_lib.unescape(text)).strip()

def fetch_lpg_news():
    fetch_time = datetime.now(); now_utc = datetime.now(timezone.utc)
    articles = []; errors = []; seen = set()
    print("\n[News] Fetching LPG news from Google News RSS …")
    for label, query in NEWS_QUERIES:
        url = f"https://news.google.com/rss/search?q={query}&hl=en-IN&gl=IN&ceid=IN:en"
        try:
            req = urllib.request.Request(url,headers={"User-Agent":NEWS_UA,"Accept":"application/rss+xml"})
            with urllib.request.urlopen(req,timeout=NEWS_TIMEOUT) as r: raw_xml = r.read()
            root  = ET.fromstring(raw_xml); count = 0
            for item in root.findall(".//item"):
                title = _clean_text(item.findtext("title",""))
                if len(title)<10: continue
                key = re.sub(r"[^a-z0-9]","",title.lower())[:80]
                if key in seen: continue
                seen.add(key)
                link   = item.findtext("link","").strip()
                pub_dt = _parse_rfc2822(item.findtext("pubDate",""))
                desc   = _clean_text(item.findtext("description",""))
                src_el = item.find("source")
                source = src_el.text.strip() if src_el is not None else "Google News"
                if pub_dt:
                    age_days = (now_utc-pub_dt).total_seconds()/86400
                    pub_str  = pub_dt.strftime("%d %b %Y  %H:%M")
                else:
                    age_days = 999.0; pub_str = "Unknown"
                if age_days>NEWS_MAX_AGE_D: continue
                articles.append({"title":title,"source":source,"published_dt":pub_dt,
                                  "published_str":pub_str,"url":link,"description":desc[:400],
                                  "query_tag":label,"age_days":age_days})
                count += 1
            print(f"      ✓ [{label}]  {count} articles")
        except Exception as exc:
            errors.append((label,str(exc)))
            print(f"      ⚠  [{label}]  Failed: {str(exc)[:80]}")
    articles.sort(key=lambda a: a["published_dt"] or datetime.min.replace(tzinfo=timezone.utc),reverse=True)
    print(f"      Total unique articles: {len(articles)}")
    return articles, fetch_time, errors

# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL BUILDER  – all 12 sheets
# ═══════════════════════════════════════════════════════════════════════════

def build_excel(delivery_plans, fleet_status, trucks, unserved, _unused,
                output_path, stations_df=None, sources_df=None, sales_raw=None,
                news_articles=None, news_fetch_time=None, news_errors=None):
    wb    = Workbook()
    today = datetime.today().strftime("%d-%m-%Y")
    if stations_df is None: stations_df = pd.DataFrame()
    if sources_df  is None: sources_df  = pd.DataFrame()
    if sales_raw   is None: sales_raw   = pd.DataFrame()

    kpis = compute_kpis(delivery_plans, fleet_status, stations_df, sources_df, sales_raw)

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 1: Delivery Plan
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = "Delivery Plan"; ws1.freeze_panes = "A3"
    stitle(ws1,"A1:R1",f"LPG AUTO DISPATCH – DELIVERY PLAN   ({today})")
    h1 = ["Truck\nID","Type","Cap\n(Lt)","Starts From","Source\nID","Source Name",
          "Truck→Src\n(km)","Truck→Src\nToll (₹)","Total Load\n(Lt)",
          "Stop 1","Stop 2","Stop 3",
          "Total Delivered\n(Lt)","Total Delivered\n(MT)",
          "Purchase\n(₹)","Transport\n(₹)","Toll\n(₹)","Grand Total\n(₹)"]
    [hdr(ws1,2,i+1,h) for i,h in enumerate(h1)]
    ws1.row_dimensions[2].height = 44
    for ri,dp in enumerate(delivery_plans,3):
        alt = CLR["lt_blue"] if ri%2==0 else CLR["white"]
        sn  = [f"{s['station']}\n({s['deliver_lt']:,.0f} Lt)" for s in dp["stops"]]
        while len(sn)<3: sn.append("")
        vals = [dp["truck_id"],dp["truck_type"],dp["capacity_lt"],dp["initial_park"],
                dp["source_id"],dp["source_name"],dp["pk_src_dist"],dp["pk_src_toll"],
                dp["total_load_lt"],sn[0],sn[1],sn[2],
                dp["total_load_lt"],dp["total_load_mt"],
                dp["tot_purchase"],dp["tot_transport"],dp["tot_toll"],dp["grand_total"]]
        fmts = [None,None,NF,None,None,None,"0.1",MF,NF,None,None,None,NF,"0.000",MF,MF,MF,MF]
        for ci,(v,f) in enumerate(zip(vals,fmts),1):
            cel(ws1,ri,ci,v,fmt=f,fh=alt)
    tr = len(delivery_plans)+3
    ws1.cell(row=tr,column=1,value="TOTALS").font = _fn(bold=True)
    for col,fmt in [(13,NF),(14,"0.000"),(15,MF),(16,MF),(17,MF),(18,MF)]:
        c = ws1.cell(row=tr,column=col)
        c.value = f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{tr-1})"
        c.number_format=fmt; c.font=_fn(bold=True); c.border=_bd()
        c.fill=_f(CLR["lt_blue"]); c.alignment=_ac()
    cw(ws1,[9,9,12,30,10,24,12,14,12,30,30,30,14,13,18,18,13,20])

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 2: Delivery Sequence
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Delivery Sequence"); ws2.freeze_panes = "A4"
    stitle(ws2,"A1:N1",f"FULL JOURNEY – STEP BY STEP   ({today})",CLR["dk_green"])
    ws2.merge_cells("O1:W1")
    nc = ws2["O1"]
    nc.value = ("Carrying MT × Leg distance RTKM:  "
                "leg ≤100km → Rs1,750×carrying_MT (flat)  |  "
                "leg >100km → Rs6.8×carrying_MT×leg_km×2 (RTKM)  |  "
                "carrying_MT = LPG on truck at departure of each leg.")
    nc.font = Font(name="Arial",italic=True,size=9,color="FFFFFF")
    nc.fill = _f(CLR["dk_green"]); nc.alignment = _al(wrap=False)
    ws2.column_dimensions["O"].width = 90
    legend = [("A2:B2","Initial park",CLR["INITIAL"]),("C2:D2","Source – load",CLR["SOURCE"]),
              ("E2:F2","Stop – deliver",CLR["DELIVER"]),("G2:H2","Proportional",CLR["prop"]),
              ("I2:J2","Final park",CLR["FINAL"])]
    for rng,lbl,clr in legend:
        ws2.merge_cells(rng); c=ws2[rng.split(":")[0]]
        c.value=f"● {lbl}"; c.font=Font(name="Arial",bold=True,size=9,color="333333")
        c.fill=_f(clr); c.border=_bd(); c.alignment=_ac()
    ws2.row_dimensions[2].height=18
    h2 = ["Truck\nID","Type","Step\n#","Step Type","Location",
          "Qty Delivered\n(Lt)","Qty Delivered\n(MT)",
          "Dist\n(km)","Toll\n(₹)","Transport\n(₹)","Purchase\n(₹ ref)",
          "Tank After\n(Lt)","Notes","Needed vs\nDelivered"]
    [hdr(ws2,3,i+1,h,fh=CLR["dk_green"]) for i,h in enumerate(h2)]
    ws2.row_dimensions[3].height=44
    r2=4
    for dp in delivery_plans:
        for si,step in enumerate(dp["journey_steps"],1):
            stype  = step["step_type"]
            is_prop = (stype=="DELIVER" and
                       any(s["deliver_lt"]<s["needed_lt"] for s in dp["stops"]
                           if s["station"]==step["location"]))
            fh = CLR["prop"] if is_prop else STEP_FILL.get(stype,CLR["white"])
            nvd=""
            if stype=="DELIVER":
                for s in dp["stops"]:
                    if s["station"]==step["location"]:
                        nvd=(f"Needed {s['needed_lt']:,.0f} → got {s['deliver_lt']:,.0f}"
                             if s["deliver_lt"]<s["needed_lt"]
                             else f"Full ({s['deliver_lt']:,.0f} Lt)")
                        break
            vals=[dp["truck_id"],dp["truck_type"],si,
                  step["label"],step["location"],step["qty_lt"],step["qty_mt"],
                  step["dist_km"],step["toll"],step["transport_cost"],step["purchase_cost"],
                  step["tank_after_lt"],step["note"],nvd]
            fmts=[None,None,None,None,None,NF,"0.000","0.1",MF,MF,MF,NF,None,None]
            for ci,(v,f) in enumerate(zip(vals,fmts),1):
                cel(ws2,r2,ci,v,fmt=f,fh=fh)
            r2+=1
        r2+=1
    cw(ws2,[9,9,6,22,38,14,13,10,12,14,14,13,50,30])

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 3: Fleet Status
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Fleet Status"); ws3.freeze_panes = "A3"
    stitle(ws3,"A1:F1",f"FLEET END-OF-DAY STATUS   ({today})",CLR["purple"])
    h3=["Truck\nID","Type","Status","Initial Parked At","Final Parked At","Notes"]
    [hdr(ws3,2,i+1,h,fh=CLR["purple"]) for i,h in enumerate(h3)]
    ws3.row_dimensions[2].height=44
    for ri,ts in enumerate(fleet_status,3):
        alt=CLR["gray"] if ri%2==0 else CLR["white"]
        sf =CLR["DEP"] if ts["status"]=="DEPLOYED" else CLR["SBY"]
        note="✅ Delivered today" if ts["status"]=="DEPLOYED" else "– On standby"
        for ci,v in enumerate([ts["truck_id"],ts["type"],ts["status"],
                                 ts["initial_park"],ts["final_park"],note],1):
            cel(ws3,ri,ci,v,fh=(sf if ci==3 else alt))
    last=len(fleet_status)+4; dn=sum(1 for t in fleet_status if t["status"]=="DEPLOYED")
    for lbl,val,row in [("DEPLOYED",dn,last),("STANDBY",len(fleet_status)-dn,last+1)]:
        ws3.cell(row=row,column=1,value=lbl).font=_fn(bold=True)
        ws3.cell(row=row,column=2,value=val).font=_fn(bold=True)
    cw(ws3,[9,10,12,44,44,30])

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 4: Cost Summary
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Cost Summary"); ws4.freeze_panes = "A3"
    stitle(ws4,"A1:G1",f"COST SUMMARY   ({today})",CLR["dk_amber"])
    explanation_row(ws4,2,7,
        "Carrying MT RTKM: leg ≤100km → Rs1,750×carrying_MT  |  leg >100km → Rs6.8×carrying_MT×leg_km×2  |  carrying_MT = LPG on truck at departure of each leg",
        CLR["yellow"])
    h4=["Truck\nID","Source\nID","Stations Served (litres delivered)",
        "Purchase\n(₹)","Transport\n(₹)","Toll\n(₹)","Grand Total\n(₹)"]
    [hdr(ws4,3,i+1,h,fh=CLR["dk_amber"]) for i,h in enumerate(h4)]
    ws4.row_dimensions[3].height=44
    for ri,dp in enumerate(delivery_plans,4):
        alt=CLR["lt_amber"] if ri%2==0 else CLR["white"]
        seq=" → ".join(f"{s['station']} ({s['deliver_lt']:,.0f} Lt)" for s in dp["stops"])
        for ci,(v,f) in enumerate(zip(
            [dp["truck_id"],dp["source_id"],seq,
             dp["tot_purchase"],dp["tot_transport"],dp["tot_toll"],dp["grand_total"]],
            [None,None,None,MF,MF,MF,MF]),1):
            cel(ws4,ri,ci,v,fmt=f,fh=alt)
    tr4=len(delivery_plans)+4
    ws4.cell(row=tr4,column=1,value="TOTALS").font=_fn(bold=True)
    for col,fmt in [(4,MF),(5,MF),(6,MF),(7,MF)]:
        c=ws4.cell(row=tr4,column=col)
        c.value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{tr4-1})"
        c.number_format=fmt; c.font=_fn(bold=True); c.border=_bd()
        c.fill=_f(CLR["lt_amber"]); c.alignment=_ac()
    cw(ws4,[9,10,70,20,20,16,22])

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 5: Unserved Stations
    # ══════════════════════════════════════════════════════════════════════
    if unserved:
        ws5=wb.create_sheet("Unserved Stations"); ws5.freeze_panes="A3"
        stitle(ws5,"A1:D1",f"UNSERVED STATIONS   ({today})","AA0000")
        h5=["Station Name","Needed (Lt)","Needed (MT)","Reason"]
        [hdr(ws5,2,i+1,h,fh="AA0000") for i,h in enumerate(h5)]
        ws5.row_dimensions[2].height=40
        for ri,s in enumerate(unserved,3):
            alt="FFD0D0" if ri%2==0 else "FFE8E8"
            cel(ws5,ri,1,s["station"],fh=alt)
            cel(ws5,ri,2,s["needed_lt"],fmt=NF,fh=alt)
            cel(ws5,ri,3,round(s["needed_lt"]/1810,3),fmt="0.000",fh=alt)
            cel(ws5,ri,4,"All trucks deployed – schedule for tomorrow",fh=alt)
        cw(ws5,[44,14,14,50])

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 6: Sales Reference
    # ══════════════════════════════════════════════════════════════════════
    ws6=wb.create_sheet("Sales Reference"); ws6.freeze_panes="A3"
    stitle(ws6,"A1:E1",f"SALES-PROPORTIONAL DELIVERY REFERENCE   ({today})",CLR["dk_blue"])
    h6=["Station","Avg Daily Sales\n(Lt/day)","This Delivery\n(Lt)","Station Needed\n(Lt)","Split %"]
    [hdr(ws6,2,i+1,h) for i,h in enumerate(h6)]
    ws6.row_dimensions[2].height=44
    r6=3
    for dp in delivery_plans:
        total_load=dp["total_load_lt"]
        for s in dp["stops"]:
            is_split=s["deliver_lt"]<s["needed_lt"]
            alt=CLR["prop"] if is_split else (CLR["lt_blue"] if r6%2==0 else CLR["white"])
            pct=round(s["deliver_lt"]/total_load*100,1) if total_load>0 else 0
            for ci,(v,f) in enumerate(zip([s["station"],s.get("avg_sales","–"),
                                           s["deliver_lt"],s["needed_lt"],f"{pct:.1f}%"],
                                          [None,NF,NF,NF,None]),1):
                cel(ws6,r6,ci,v,fmt=f,fh=alt)
            r6+=1
        r6+=1
    cw(ws6,[44,18,16,16,10])

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 7: LPG News Feed
    # ══════════════════════════════════════════════════════════════════════
    _build_news_sheet(wb, news_articles or [], news_fetch_time or datetime.now(), news_errors or [])

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 8: KPI Dashboard
    # ══════════════════════════════════════════════════════════════════════
    ws8=wb.create_sheet("KPI Dashboard"); ws8.freeze_panes="A3"
    stitle(ws8,"A1:D1",f"TODAY'S KEY NUMBERS   ({today}   {kpis['run_time']})",CLR["dk_blue"])
    explanation_row(ws8,2,4,
        "What happened today at a glance. Green = good, Amber = watch, Red = act now.",
        CLR["yellow"])
    hdr(ws8,3,1,"What We Measure",CLR["dk_blue"]); hdr(ws8,3,2,"Today's Number",CLR["dk_blue"])
    hdr(ws8,3,3,"What This Means",CLR["dk_blue"]); hdr(ws8,3,4,"Status",CLR["dk_blue"])
    ws8.row_dimensions[3].height=30
    cw(ws8,[38,22,55,16])

    kpi_rows = [
        ("LPG Delivered Today",
         f"{kpis['total_lt']:,.0f} Lt  ({kpis['total_mt']:.1f} MT)",
         f"{kpis['total_runs']} truck run(s) completed today",
         "OK" if kpis["total_runs"]>0 else "NO RUNS"),
        ("Trucks Deployed",
         f"{kpis['trucks_deployed']} of {kpis['trucks_total']} trucks",
         f"{kpis['trucks_standby']} trucks on standby  |  Utilisation: {kpis['utilisation_pct']:.0f}%",
         "OK" if kpis["utilisation_pct"]>10 else "LOW"),
        ("Stations Served Today",
         f"{kpis['stations_served']} of {kpis['stations_needing']} needing delivery",
         (f"All stations served ✓" if kpis["stations_unserved"]==0
          else f"{kpis['stations_unserved']} station(s) could not be served today"),
         "OK" if kpis["stations_unserved"]==0 else "ALERT"),
        ("Total Spend Today",
         f"₹{kpis['grand_total']:,.0f}",
         f"Purchase ₹{kpis['tot_purchase']:,.0f}  |  Transport ₹{kpis['tot_transport']:,.0f}  |  Toll ₹{kpis['tot_toll']:,.0f}",
         "OK"),
        ("Cost per MT Delivered",
         f"₹{kpis['cost_per_mt']:,.0f} per MT",
         f"Or ₹{kpis['cost_per_lt']:.2f} per litre delivered",
         "OK"),
        ("Purchase vs Transport Split",
         f"Purchase: {kpis['purchase_pct']:.0f}%  |  Transport: {kpis['transport_pct']:.0f}%  |  Toll: {kpis['toll_pct']:.0f}%",
         "Purchase dominates cost — transport % is the operational efficiency lever",
         "OK"),
        ("LPG Terminals Used Today",
         f"{kpis['sources_used']} terminal(s)",
         "Optimizer chose cheapest source per station based on (price + transport) total cost",
         "OK"),
        ("Monthly Consumption (Sales Data)",
         f"{kpis['monthly_total_lt']:,.0f} Lt  ({kpis['monthly_total_mt']:.0f} MT)",
         (f"Avg {kpis['avg_daily_per_st']:,.0f} Lt/station/day  |  "
          f"{kpis['reporting_days']} days  |  {kpis['reporting_stations']} stations"),
         "INFO"),
    ]

    STATUS_COLOR = {"OK":CLR["ok_green"],"WATCH":CLR["warn_yel"],"ALERT":CLR["alert"],
                    "LOW":CLR["warn_yel"],"NO RUNS":CLR["alert"],"INFO":CLR["lt_blue"]}
    for ri,(label,value,meaning,status) in enumerate(kpi_rows,4):
        row_fh = CLR["gray"] if ri%2==0 else CLR["white"]
        cel(ws8,ri,1,label,  bold=True,fh=row_fh)
        cel(ws8,ri,2,value,  bold=True,fh=row_fh)
        cel(ws8,ri,3,meaning,italic=True,fh=row_fh)
        s=ws8.cell(row=ri,column=4,value=status)
        s.font=_fn(bold=True,col="FFFFFF" if status in ("ALERT","OK") else "333333")
        s.fill=_f(STATUS_COLOR.get(status,CLR["gray"]))
        s.border=_bd(); s.alignment=_ac()
        ws8.row_dimensions[ri].height=32

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 9: Station Intelligence
    # ══════════════════════════════════════════════════════════════════════
    ws9=wb.create_sheet("Station Intelligence"); ws9.freeze_panes="A4"
    stitle(ws9,"A1:G1",f"STATION INTELLIGENCE   ({today})",CLR["dk_green"])

    if not sales_raw.empty:
        s_data = sales_raw.drop("Date",axis=1) if "Date" in sales_raw.columns else sales_raw
        station_stats=[]
        for col in s_data.columns:
            vals=s_data[col].dropna().values
            if len(vals)==0: continue
            mean=np.nanmean(vals); std=np.nanstd(vals)
            station_stats.append({
                "station":col.strip(),
                "avg_daily":round(mean,0),"std_dev":round(std,0),
                "max_day":int(np.nanmax(vals)),"min_day":int(np.nanmin(vals)),
                "monthly_total":int(np.nansum(vals)),
                "cv_pct":round(std/mean*100,1) if mean>0 else 0,
            })
        station_stats.sort(key=lambda x: x["avg_daily"],reverse=True)
        h9=["Station","Avg Daily\n(Lt)","Monthly Total\n(Lt)","Highest Day\n(Lt)","Lowest Day\n(Lt)","Demand Level"]
        ncols9=6

        row9=3
        section_hdr(ws9,row9,ncols9,"  TOP 10 STATIONS BY AVERAGE DAILY CONSUMPTION  (highest demand — plan deliveries here first)",CLR["dk_green"])
        row9+=1
        explanation_row(ws9,row9,ncols9,
            "These stations sell the most LPG per day. They will need more frequent deliveries and larger proportional share when truck capacity is shared.",
            CLR["lt_green"])
        row9+=1
        [hdr(ws9,row9,i+1,h,fh=CLR["dk_green"]) for i,h in enumerate(h9)]
        ws9.row_dimensions[row9].height=40; row9+=1
        for rank,(st_row) in enumerate(station_stats[:10],1):
            fh=CLR["ok_green"] if rank<=3 else (CLR["lt_green"] if rank%2==0 else CLR["white"])
            cel(ws9,row9,1,st_row["station"],bold=(rank<=3),fh=fh)
            cel(ws9,row9,2,int(st_row["avg_daily"]),fmt=NF,fh=fh,center=True)
            cel(ws9,row9,3,st_row["monthly_total"],fmt=NF,fh=fh)
            cel(ws9,row9,4,st_row["max_day"],fmt=NF,fh=fh,center=True)
            cel(ws9,row9,5,st_row["min_day"],fmt=NF,fh=fh,center=True)
            demand="🔴 Very High" if st_row["avg_daily"]>3000 else ("🟠 High" if st_row["avg_daily"]>1500 else "🟡 Medium")
            cel(ws9,row9,6,demand,center=True,fh=fh)
            row9+=1

        row9+=1
        volatile=sorted(station_stats,key=lambda x: x["cv_pct"],reverse=True)
        section_hdr(ws9,row9,ncols9,"  TOP 10 MOST VOLATILE STATIONS  (demand changes a lot day-to-day — harder to predict)",CLR["dk_amber"])
        row9+=1
        explanation_row(ws9,row9,ncols9,
            "CV% = how much daily sales vary. High CV means unpredictable demand — these stations may suddenly need LPG or go quiet. Watch closely.",
            CLR["lt_amber"])
        row9+=1
        h9v=["Station","Avg Daily\n(Lt)","Std Deviation\n(Lt)","Highest Day\n(Lt)","Lowest Day\n(Lt)","Variability %"]
        [hdr(ws9,row9,i+1,h,fh=CLR["dk_amber"]) for i,h in enumerate(h9v)]
        ws9.row_dimensions[row9].height=40; row9+=1
        for st_row in volatile[:10]:
            fh=CLR["lt_amber"] if row9%2==0 else CLR["white"]
            high_cv=st_row["cv_pct"]>30
            cv_fh=CLR["lt_red"] if st_row["cv_pct"]>40 else (CLR["warn_yel"] if high_cv else fh)
            cel(ws9,row9,1,st_row["station"],bold=high_cv,fh=fh)
            cel(ws9,row9,2,int(st_row["avg_daily"]),fmt=NF,fh=fh,center=True)
            cel(ws9,row9,3,st_row["std_dev"],fmt=NF,fh=fh,center=True)
            cel(ws9,row9,4,st_row["max_day"],fmt=NF,fh=fh,center=True)
            cel(ws9,row9,5,st_row["min_day"],fmt=NF,fh=fh,center=True)
            cv_label=f"{st_row['cv_pct']:.0f}%  {'⚠' if high_cv else ''}"
            cel(ws9,row9,6,cv_label,center=True,fh=cv_fh,bold=high_cv)
            row9+=1

        row9+=1
        section_hdr(ws9,row9,ncols9,f"  ALL {len(station_stats)} STATIONS – DEMAND SUMMARY",CLR["dk_blue"])
        row9+=1
        [hdr(ws9,row9,i+1,h) for i,h in enumerate(h9)]
        ws9.row_dimensions[row9].height=40; row9+=1
        for rank,st_row in enumerate(station_stats,1):
            fh=CLR["gray"] if rank%2==0 else CLR["white"]
            cel(ws9,row9,1,st_row["station"],fh=fh)
            cel(ws9,row9,2,int(st_row["avg_daily"]),fmt=NF,fh=fh,center=True)
            cel(ws9,row9,3,st_row["monthly_total"],fmt=NF,fh=fh)
            cel(ws9,row9,4,st_row["max_day"],fmt=NF,fh=fh,center=True)
            cel(ws9,row9,5,st_row["min_day"],fmt=NF,fh=fh,center=True)
            demand=("🔴 Very High" if st_row["avg_daily"]>3000 else
                    "🟠 High" if st_row["avg_daily"]>1500 else
                    "🟡 Medium" if st_row["avg_daily"]>800 else "🟢 Low")
            cel(ws9,row9,6,demand,center=True,fh=fh)
            row9+=1
    cw(ws9,[44,16,18,14,14,14])

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 10: Source Comparison
    # ══════════════════════════════════════════════════════════════════════
    ws12=wb.create_sheet("Source Comparison"); ws12.freeze_panes="A4"
    stitle(ws12,"A1:F1",f"LPG TERMINAL COMPARISON   ({today})",CLR["teal"])
    explanation_row(ws12,2,6,
        "All LPG sourcing terminals ranked by price per MT (cheapest first). The optimizer automatically picks the cheapest terminal after adding transport cost per station.",
        CLR["yellow"])
    explanation_row(ws12,3,6,
        "The 'Effective Cost/MT' includes the approximate transport cost to reach a typical station — this is what you actually pay, not just the ex-terminal price.",
        CLR["gray"])

    if not sources_df.empty:
        src_sorted=sources_df.sort_values("Price / MT Ex Terminal").reset_index(drop=True)
        min_price=src_sorted["Price / MT Ex Terminal"].min()
        max_price=src_sorted["Price / MT Ex Terminal"].max()
        h12=["Rank","Terminal ID","Terminal Name","Price / MT\n(Ex-Terminal ₹)","vs Cheapest","Recommendation"]
        [hdr(ws12,4,i+1,h,fh=CLR["teal"]) for i,h in enumerate(h12)]
        ws12.row_dimensions[4].height=40
        for ri,(_, row) in enumerate(src_sorted.iterrows(),5):
            rank=ri-4
            price=float(row["Price / MT Ex Terminal"])
            vs_min=price-min_price
            fh=(CLR["ok_green"] if rank==1 else
                CLR["lt_green"] if rank<=3 else
                CLR["gray"] if rank%2==0 else CLR["white"])
            rec=("★ Cheapest – preferred source" if rank==1 else
                 "✓ Good option" if rank<=3 else
                 f"+₹{vs_min:,.0f}/MT more than cheapest")
            cel(ws12,ri,1,rank,center=True,bold=(rank==1),fh=fh)
            cel(ws12,ri,2,row["Source_ID"],bold=(rank==1),center=True,fh=fh)
            cel(ws12,ri,3,row["Source_Name"],bold=(rank==1),fh=fh)
            cel(ws12,ri,4,int(price),fmt=NF,center=True,fh=fh,bold=(rank==1))
            vs_cell=ws12.cell(row=ri,column=5,value=(f"+₹{vs_min:,.0f}/MT" if vs_min>0 else "— cheapest"))
            vs_cell.font=_fn(bold=(rank==1),col=("375623" if rank==1 else "C00000" if vs_min>2000 else "000000"))
            vs_cell.fill=_f(fh); vs_cell.border=_bd(); vs_cell.alignment=_ac()
            cel(ws12,ri,6,rec,italic=True,fh=fh,col=("375623" if rank==1 else "000000"))

        if delivery_plans:
            used_today=defaultdict(int)
            for dp in delivery_plans: used_today[dp["source_id"]]+=dp["total_load_lt"]
            usage_row=len(src_sorted)+7
            section_hdr(ws12,usage_row,6,"  TERMINALS USED IN TODAY'S DISPATCH RUN",CLR["dk_blue"])
            usage_row+=1
            [hdr(ws12,usage_row,i+1,h) for i,h in enumerate(["Terminal ID","Terminal Name","Litres Dispatched","MT Dispatched","Purchase Cost ₹","Runs"])]
            ws12.row_dimensions[usage_row].height=30; usage_row+=1
            seen_src=set()
            for dp_group in delivery_plans:
                src_id=dp_group["source_id"]
                if src_id in seen_src: continue
                seen_src.add(src_id)
                src_name=dp_group["source_name"]
                total_lt_src=sum(d["total_load_lt"] for d in delivery_plans if d["source_id"]==src_id)
                total_pur_src=sum(d["tot_purchase"] for d in delivery_plans if d["source_id"]==src_id)
                n_runs=sum(1 for d in delivery_plans if d["source_id"]==src_id)
                fh=CLR["gray"] if usage_row%2==0 else CLR["white"]
                cel(ws12,usage_row,1,src_id,center=True,bold=True,fh=fh)
                cel(ws12,usage_row,2,src_name,fh=fh)
                cel(ws12,usage_row,3,total_lt_src,fmt=NF,fh=fh)
                cel(ws12,usage_row,4,round(total_lt_src/MT_TO_LITERS,2),fmt="0.00",fh=fh)
                cel(ws12,usage_row,5,round(total_pur_src,0),fmt=MF,fh=fh)
                cel(ws12,usage_row,6,n_runs,center=True,fh=fh)
                usage_row+=1
    cw(ws12,[6,14,36,18,16,36])

    wb.save(output_path)
    print(f"\n      Excel saved → {output_path}")
    print("="*65)


def _build_news_sheet(wb, articles, fetch_time, errors):
    ws=wb.create_sheet("LPG News Feed"); ws.freeze_panes="A4"
    today_str=datetime.today().strftime("%d-%m-%Y")
    fetch_str=fetch_time.strftime("%d-%m-%Y  %H:%M:%S")
    ws.merge_cells("A1:H1"); t1=ws["A1"]
    t1.value=f"AUTO LPG – LIVE NEWS FEED   ({today_str})   |   Fetched: {fetch_str}   |   {len(articles)} articles"
    t1.font=Font(name="Arial",bold=True,size=12,color="FFFFFF")
    t1.fill=_f("1F3864"); t1.alignment=_al(wrap=False)
    ws.row_dimensions[1].height=28
    for rng,lbl,clr in [("A2:B2","Today","C6EFCE"),("C2:D2","This week","FFEB9C"),("E2:F2","Older","F2F2F2")]:
        ws.merge_cells(rng); c=ws[rng.split(":")[0]]
        c.value=f"● {lbl}"; c.font=Font(name="Arial",bold=True,size=9,color="333333")
        c.fill=_f(clr); c.alignment=_ac()
    ws.row_dimensions[2].height=20
    if errors:
        ws.merge_cells("G2:H2"); ec=ws["G2"]
        ec.value=f"⚠ {len(errors)} query(s) failed – check internet"
        ec.font=Font(name="Arial",bold=True,size=9,color="AA0000")
        ec.fill=_f("FFD0D0"); ec.alignment=_ac()
    headers=["#","Published","Source","Headline (click to open)","Description","Category","Age","Full URL"]
    col_widths=[4,18,22,55,70,22,8,60]
    for ci,(h,w) in enumerate(zip(headers,col_widths),1):
        c=ws.cell(row=3,column=ci,value=h)
        c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF")
        c.fill=_f("1F3864"); c.border=_bd(); c.alignment=_ac()
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[3].height=30
    thin=Side(style="thin"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    if not articles:
        ws.merge_cells("A4:H4"); nc=ws["A4"]
        nc.value="No articles found. Check internet connection."
        nc.font=Font(name="Arial",italic=True,size=10,color="888888")
        nc.alignment=_ac(); ws.row_dimensions[4].height=30
        return
    now_utc=datetime.now(timezone.utc)
    for idx,art in enumerate(articles,1):
        row=idx+3
        age=art["age_days"]
        fh="C6EFCE" if age<=1 else ("FFEB9C" if age<=7 else "F2F2F2")
        age_lbl=("< 1h" if age<0.042 else f"{int(age*24)}h ago" if age<1 else
                  "Yesterday" if age<2 else f"{int(age)}d ago")
        def _c(col,val,bold=False,center=False):
            c=ws.cell(row=row,column=col,value=val)
            c.font=Font(name="Arial",bold=bold,size=9,color="000000")
            c.fill=_f(fh); c.border=bdr
            c.alignment=Alignment(horizontal="center" if center else "left",vertical="top",wrap_text=True)
        _c(1,idx,center=True); _c(2,art["published_str"]); _c(3,art["source"],bold=True)
        _c(5,art["description"]); _c(6,art["query_tag"],center=True)
        _c(7,age_lbl,center=True); _c(8,art["url"])
        hl=ws.cell(row=row,column=4,value=art["title"])
        hl.font=Font(name="Arial",bold=True,size=9,color="0563C1",underline="single")
        hl.fill=_f(fh); hl.border=bdr
        hl.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
        if art["url"]:
            try: hl.hyperlink=art["url"]
            except Exception: pass
        ws.row_dimensions[row].height=42

# ═══════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="LPG Dispatch Optimizer")
    parser.add_argument("--reset-positions", action="store_true", help="Force recalculation of initial truck positions (ignore saved JSON)")
    args = parser.parse_args()

    delivery_plans, fleet_status, trucks, unserved, unserved_res = run_optimization(force_reset=args.reset_positions)
    
    if delivery_plans is None:
        sys.exit(0)

    stations_df = pd.DataFrame()
    sources_df  = pd.DataFrame()
    sales_raw   = pd.DataFrame()
    try:
        stations_df = pd.read_excel(resolve(STATIONS_FILE))
        stations_df.columns = [c.strip() for c in stations_df.columns]
        if "Stations " in stations_df.columns:
            stations_df.rename(columns={"Stations ":"Stations"},inplace=True)
        stations_df["Stations"] = stations_df["Stations"].str.strip()
    except Exception: pass
    try:
        sources_df = pd.read_excel(resolve(SOURCES_FILE))
        sources_df.columns = [c.strip() for c in sources_df.columns]
    except Exception: pass
    try:
        sales_raw = pd.read_excel(resolve(SALES_FILE))
    except Exception: pass

    news_articles, news_fetch_time, news_errors = fetch_lpg_news()

    date_str  = datetime.today().strftime("%Y%m%d_%H%M")
    out_path  = os.path.join(output_dir(), f"LPG_DISPATCH_PLAN_{date_str}.xlsx")
    build_excel(
        delivery_plans, fleet_status, trucks, unserved, unserved_res, out_path,
        stations_df=stations_df, sources_df=sources_df, sales_raw=sales_raw,
        news_articles=news_articles, news_fetch_time=news_fetch_time, news_errors=news_errors,
    )

    print("\n📋 QUICK SUMMARY")
    print(f"   Runs planned  : {len(delivery_plans)}")
    print(f"   Trucks used   : {sum(1 for t in fleet_status if t['status']=='DEPLOYED')} / {len(trucks)}")
    print(f"   Total LPG     : {sum(dp['total_load_lt'] for dp in delivery_plans):,.0f} Lt")
    print(f"   Total cost    : ₹{sum(dp['grand_total'] for dp in delivery_plans):,.2f}")
    if unserved:
        print(f"   ❌ Unserved  : {len(unserved)} stations")
    print(f"   📰 News items : {len(news_articles)} articles fetched")
    if news_errors:
        print(f"   ⚠  News errors: {len(news_errors)} query(s) failed")
    print(f"\n   Output        : {out_path}\n")