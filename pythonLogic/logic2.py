"""
=============================================================================
  LPG AUTO DISPATCH OPTIMIZER  –  Full Standalone Script
  Run:  python3 FULL_LPG_OPTIMIZATION.py
  Output: LPG_DISPATCH_PLAN_<date>.xlsx

  PERSISTENT STATE:
    truck_positions.json  – saved after every run, loaded at start of next run
    • First run ever   → trucks placed by max-dispersion across all 81 stations
    • Every later run  → trucks start exactly where they parked last time
    • Reset positions  → delete truck_positions.json and re-run

  RELOAD LOGIC:
    If a truck's capacity is less than the total LPG needed for all its stops,
    it must return to the source between deliveries to reload.
    Each reload adds:
      • Extra transport cost  (station → source leg)
      • Extra toll            (on that leg)
      • Extra purchase cost   (for the reloaded LPG)
    The Delivery Sequence sheet shows every step including 🔄 RELOAD legs.

  FULL JOURNEY SEQUENCE per truck (shown in Delivery Sequence sheet):
    🚚 INITIAL PARK
    ⛽ SOURCE – LOAD
    📍 STOP 1 – DELIVER
    🔄 RELOAD at Source   ← inserted automatically when capacity exceeded
    📍 STOP 2 – DELIVER
    🏁 FINAL PARK
=============================================================================
"""

import pandas as pd
import requests
import math, time, os, sys, json
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_unserved_stations = []   # stations that couldn't be served (no truck available)

# ═══════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════

STATIONS_FILE   = "clean_stationss.xlsx"
SOURCES_FILE    = "sources.xlsx"
POSITIONS_FILE  = "truck_positions.json"
GOOGLE_API_KEY  = "AIzaSyA8oVRSa2W2IzX9hg4vnaKM6hwkGRGnsP4"

FLEET = [
    {"type": "12MT", "count": 23, "capacity_mt": 12, "capacity_lt": 12 * 1810},
    {"type": "7MT",  "count":  7, "capacity_mt":  7, "capacity_lt":  7 * 1810},
]

MT_TO_LITERS        = 1810
TRANSPORT_FLAT      = 1750   # Rs/MT  dist < 100 km
TRANSPORT_PER_KM    = 6.8    # Rs/MT/RTKM  dist >= 100 km
MAX_STOPS_PER_TRUCK = 3
MAX_GROUPING_KM     = 80

# ── Parallelism ──────────────────────────────────────────────────────────
MAX_WORKERS         = 8    # concurrent API threads
SOURCE_PREFETCH_N   = 6    # top-N sources by approx total cost (was 4 by distance)
API_RATE_SLEEP      = 0.0  # no sleep needed – parallel requests are naturally throttled

# ── Time estimation constants ────────────────────────────────────────────
AVG_SPEED_KMH       = 40    # loaded truck speed on Tamil Nadu roads (km/h)
UNLOAD_MIN_PER_STOP = 30    # minutes to unload LPG at each station
LOAD_MIN_AT_SOURCE  = 45    # minutes to load LPG at source
WORK_DAY_HOURS      = 8     # working hours per day
WORK_DAY_MIN        = WORK_DAY_HOURS * 60   # 480 minutes

# ═══════════════════════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════════════════════

def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0; r = math.radians
    dlat = r(lat2-lat1); dlon = r(lon2-lon1)
    a = math.sin(dlat/2)**2 + math.cos(r(lat1))*math.cos(r(lat2))*math.sin(dlon/2)**2
    return 2*R*math.asin(math.sqrt(a))

def parse_coords(s):
    try:
        p = str(s).split(',')
        return float(p[0].strip()), float(p[1].strip())
    except (ValueError, IndexError):
        raise ValueError(f"Invalid coordinates: {s!r}")

def transport_cost_calc(dist_km, qty_mt):
    # <= 100 so flat rate wins at exactly 100 km (RTKM would be cheaper, flat is correct)
    if dist_km <= 100:
        return TRANSPORT_FLAT * qty_mt
    return TRANSPORT_PER_KM * qty_mt * dist_km * 2   # RTKM = round-trip

def transport_cost_empty(dist_km):
    """Deadhead cost for empty truck going to source. Minimum 1 MT charge."""
    return transport_cost_calc(dist_km, 1.0)

# ─── Google Routes API ───────────────────────────────────────────────────────

_route_cache = {}

def _rkey(olat, olon, dlat, dlon):
    return (round(olat,4), round(olon,4), round(dlat,4), round(dlon,4))

def get_road_info(olat, olon, dlat, dlon):
    """Returns (road_dist_km, toll_rs). Falls back to haversine×1.3 if API fails."""
    key = (round(olat,4), round(olon,4), round(dlat,4), round(dlon,4))
    if key in _route_cache:
        return _route_cache[key]
    try:
        resp = requests.post(
            "https://routes.googleapis.com/directions/v2:computeRoutes",
            headers={
                "Content-Type":   "application/json",
                "X-Goog-Api-Key": GOOGLE_API_KEY,
                "X-Goog-FieldMask": (
                    "routes.distanceMeters,routes.duration,"
                    "routes.travelAdvisory.tollInfo,routes.description"
                ),
            },
            json={
                "origin":      {"location": {"latLng": {"latitude": olat, "longitude": olon}}},
                "destination": {"location": {"latLng": {"latitude": dlat, "longitude": dlon}}},
                "travelMode":  "DRIVE",
                "computeAlternativeRoutes": True,
                "extraComputations": ["TOLLS"],
                "routeModifiers": {"vehicleInfo": {"emissionType": "DIESEL"}},
            },
            timeout=15,
        )
        data = resp.json()
        if "routes" in data and data["routes"]:
            # Among alternatives, pick the one with lowest (dist + toll×weight)
            def route_score(r):
                dist = r["distanceMeters"] / 1000
                toll = sum(float(p.get("units",0)) + float(p.get("nanos",0))/1e9
                           for p in r.get("travelAdvisory",{}).get("tollInfo",{})
                              .get("estimatedPrice",[]))
                return dist + toll * 50   # weight toll heavily
            best = min(data["routes"], key=route_score)
            dist_km = best["distanceMeters"] / 1000.0
            toll    = sum(float(p.get("units",0)) + float(p.get("nanos",0))/1e9
                         for p in best.get("travelAdvisory",{}).get("tollInfo",{})
                            .get("estimatedPrice",[]))
            _route_cache[key] = (round(dist_km,1), round(toll,2))
            return round(dist_km,1), round(toll,2)
    except Exception:
        pass
    dist = round(haversine(olat,olon,dlat,dlon)*1.3, 1)
    _route_cache[key] = (dist, 0.0)
    return dist, 0.0

# ═══════════════════════════════════════════════════════════════════════════
#  DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════

def output_dir():
    d = "/mnt/user-data/outputs"
    if os.path.isdir(d): return d
    return os.path.dirname(os.path.abspath(__file__))

def resolve(filename):
    for p in [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), filename),
        os.path.join(output_dir(), filename),
        f"/mnt/user-data/uploads/{filename}",
    ]:
        if os.path.exists(p): return p
    raise FileNotFoundError(filename)

def load_data():
    st  = pd.read_excel(resolve(STATIONS_FILE))
    src = pd.read_excel(resolve(SOURCES_FILE))
    for df in [st, src]: df.columns = [c.strip() for c in df.columns]
    if "Stations "  in st.columns:  st.rename(columns={"Stations ":  "Stations"}, inplace=True)
    if "Source_ID " in src.columns: src.rename(columns={"Source_ID ": "Source_ID"}, inplace=True)
    st["Stations"] = st["Stations"].str.strip()

    # ── Guard: skip rows with bad/missing coordinates ─────────────────────
    def safe_parse(s):
        try: return parse_coords(s)
        except: return (None, None)

    st[["lat","lon"]]  = pd.DataFrame(st["Coordinates"].map(safe_parse).tolist(), index=st.index)
    src[["lat","lon"]] = pd.DataFrame(src["Coordinates"].map(safe_parse).tolist(), index=src.index)

    bad_st  = st[st["lat"].isna()]
    bad_src = src[src["lat"].isna()]
    if not bad_st.empty:
        print(f"      ⚠  Skipping {len(bad_st)} stations with bad coordinates: "
              f"{bad_st['Stations'].tolist()}")
    if not bad_src.empty:
        print(f"      ⚠  Skipping {len(bad_src)} sources with bad coordinates: "
              f"{bad_src['Source_Name'].tolist()}")

    st  = st.dropna(subset=["lat","lon"])
    src = src.dropna(subset=["lat","lon"])

    # ── Guard: skip stations with Now=NO but Usable Lt <= 0 ──────────────
    bad_lt = st[(st["Now"].str.strip().str.upper() == "NO") & (st["Usable Lt"] <= 0)]
    if not bad_lt.empty:
        print(f"      ⚠  Skipping {len(bad_lt)} stations with Now=NO but Usable Lt=0: "
              f"{bad_lt['Stations'].tolist()}")
    st = st[~((st["Now"].str.strip().str.upper() == "NO") & (st["Usable Lt"] <= 0))]

    # ── Guard: deduplicate stations (same name appearing twice) ──────────
    dupes = st[st["Stations"].duplicated(keep=False)]
    if not dupes.empty:
        print(f"      ⚠  Duplicate station names found – keeping first occurrence: "
              f"{dupes['Stations'].unique().tolist()}")
    st = st.drop_duplicates(subset=["Stations"], keep="first")

    # ── Guard: flag stations needing more than any truck can carry ─────────
    max_cap = max(ft["capacity_lt"] for ft in [
        {"type": "12MT", "capacity_lt": 12 * 1810},
        {"type": "7MT",  "capacity_lt":  7 * 1810},
    ])
    impossible = st[(st["Now"].str.strip().str.upper() == "NO") & (st["Usable Lt"] > max_cap)]
    if not impossible.empty:
        print(f"      ⚠  Stations needing more than max truck capacity ({max_cap:,} Lt):")
        for _, row in impossible.iterrows():
            print(f"         {row['Stations']}: {row['Usable Lt']:,.0f} Lt "
                  f"→ will require multiple reload trips")

    return st, src

# ═══════════════════════════════════════════════════════════════════════════
#  PERSISTENT TRUCK POSITIONS
# ═══════════════════════════════════════════════════════════════════════════

def pos_file():
    return os.path.join(output_dir(), POSITIONS_FILE)

def load_positions():
    p = pos_file()
    if os.path.exists(p):
        with open(p) as f: data = json.load(f)
        print(f"      ✓ Loaded saved positions ({len(data)} trucks)  ← {p}")
        return data
    print(f"      ℹ  No saved positions – first run, using max-dispersion")
    return None

def save_positions(trucks):
    data = {t["truck_id"]: {"station": t["parked_station"],
                             "lat": t["parked_lat"], "lon": t["parked_lon"],
                             "type": t["type"]} for t in trucks}
    with open(pos_file(), "w") as f: json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"      ✓ Positions saved  → {pos_file()}")

def dispersion_indices(stations, n):
    coords = list(zip(stations["lat"], stations["lon"]))
    chosen = [0]
    while len(chosen) < n:
        best_i, best_d = -1, -1.0
        for i in range(len(coords)):
            if i in chosen: continue
            md = min(haversine(coords[i][0],coords[i][1],coords[c][0],coords[c][1]) for c in chosen)
            if md > best_d: best_d, best_i = md, i
        chosen.append(best_i)
    return chosen

def build_fleet(stations):
    saved = load_positions()
    trucks = []
    num = 1
    for ft in FLEET:
        for _ in range(ft["count"]):
            trucks.append({"truck_id": f"T{num:02d}", "type": ft["type"],
                           "capacity_mt": ft["capacity_mt"], "capacity_lt": ft["capacity_lt"],
                           "parked_station": None, "parked_lat": None, "parked_lon": None})
            num += 1
    if saved:
        for t in trucks:
            if t["truck_id"] in saved:
                p = saved[t["truck_id"]]
                t["parked_station"] = p["station"]
                t["parked_lat"]     = p["lat"]
                t["parked_lon"]     = p["lon"]
            else:
                coords = list(zip(stations["lat"], stations["lon"]))
                used   = {(round(v["lat"],4),round(v["lon"],4)) for v in saved.values()}
                bi, bd = 0, -1.0
                for i,(la,lo) in enumerate(coords):
                    d = min(haversine(la,lo,u[0],u[1]) for u in used) if used else 999.0
                    if d > bd: bd, bi = d, i
                row = stations.iloc[bi]
                t["parked_station"] = row["Stations"]
                t["parked_lat"]     = float(row["lat"])
                t["parked_lon"]     = float(row["lon"])
                used.add((round(row["lat"],4),round(row["lon"],4)))
        print("      ↩  Trucks restored from previous run positions")
    else:
        for t, idx in zip(trucks, dispersion_indices(stations, len(trucks))):
            row = stations.iloc[idx]
            t["parked_station"] = row["Stations"]
            t["parked_lat"]     = float(row["lat"])
            t["parked_lon"]     = float(row["lon"])
        print(f"      🆕 First run – trucks spread across {len(stations)} stations")
    return trucks

# ═══════════════════════════════════════════════════════════════════════════
#  BALANCED PARTITION  (4→2+2, 6→2+2+2, 5→3+2, 7→3+2+2)
# ═══════════════════════════════════════════════════════════════════════════

def balanced_partition(group, max_stops, max_km):
    if not group: return []
    n = len(group)
    ideal = max_stops
    for s in range(2, max_stops+1):
        if n % s == 0: ideal = s; break
    remaining, runs = group[:], []
    while remaining:
        run = [remaining.pop(0)]
        target = ideal if len(remaining)+1 >= ideal else max_stops
        while len(run) < target and remaining:
            last = run[-1]
            ni = min(range(len(remaining)),
                     key=lambda i: haversine(last["station_lat"],last["station_lon"],
                                             remaining[i]["station_lat"],remaining[i]["station_lon"]))
            if haversine(last["station_lat"],last["station_lon"],
                         remaining[ni]["station_lat"],remaining[ni]["station_lon"]) <= max_km:
                run.append(remaining.pop(ni))
            else:
                break
        if len(run) == max_stops and len(remaining) == 1:
            remaining.insert(0, run.pop())
        runs.append(run)
    return runs

# ═══════════════════════════════════════════════════════════════════════════
#  RELOAD-AWARE JOURNEY BUILDER
#  Builds the complete step-by-step journey for one truck run.
#  Detects when tank is insufficient for the next stop and inserts a
#  RELOAD leg (back to source) with its additional costs.
#
#  Returns:
#    journey_steps  – list of step dicts (one row each in Delivery Sequence)
#    cost_summary   – dict of totals for the Delivery Plan summary row
# ═══════════════════════════════════════════════════════════════════════════

def build_journey(truck, start_pos_dict, src_id, src_name, src_lat, src_lon,
                  ordered_stops, price_mt):
    """
    truck          – truck dict  (truck_id, type, capacity_lt, …)
    start_pos_dict – {"station", "lat", "lon"}  where truck is RIGHT NOW
    ordered_stops  – list of station dicts with needed_lt / station_lat / etc.
    price_mt       – source price per MT
    """
    cap_lt = truck["capacity_lt"]
    steps  = []

    # Running totals
    tot_purchase  = 0.0
    tot_transport = 0.0
    tot_toll      = 0.0
    n_reloads     = 0
    cum           = 0.0   # cumulative cost (for display)

    total_lt = sum(s["needed_lt"] for s in ordered_stops)

    # ── Step 0: INITIAL PARK ──────────────────────────────────────────────
    steps.append({
        "step_type":      "INITIAL_PARK",
        "label":          "🚚 INITIAL PARK",
        "location":       start_pos_dict["station"],
        "qty_lt":         None, "qty_mt":         None,
        "dist_km":        None, "toll":            None,
        "transport_cost": None, "purchase_cost":   None,
        "leg_cost":       None, "cum_cost":        None,
        "tank_after_lt":  None,
        "note":           "Truck starting position",
    })

    # ── Step 1: Drive to source, load first batch ─────────────────────────
    pk_dist, pk_toll = get_road_info(
        start_pos_dict["lat"], start_pos_dict["lon"], src_lat, src_lon)

    # Calculate first load: fill up as many stops as capacity allows
    remaining_stops = ordered_stops[:]
    first_load_lt   = 0.0
    for stop in remaining_stops:
        if first_load_lt + stop["needed_lt"] <= cap_lt:
            first_load_lt += stop["needed_lt"]
        else:
            break
    first_load_lt = min(first_load_lt, cap_lt)
    first_load_mt = first_load_lt / MT_TO_LITERS

    pk_tc       = transport_cost_empty(pk_dist)  # empty truck going to source
    first_purch = first_load_mt * price_mt
    tot_purchase  += first_purch
    tot_transport += pk_tc
    tot_toll      += pk_toll
    cum           += pk_tc + pk_toll + first_purch

    tank_lt = first_load_lt   # current tank level

    steps.append({
        "step_type":      "LOAD",
        "label":          "⛽ SOURCE – LOAD",
        "location":       f"{src_name}  ({src_id})",
        "qty_lt":         first_load_lt,
        "qty_mt":         round(first_load_mt, 3),
        "dist_km":        pk_dist,
        "toll":           pk_toll,
        "transport_cost": round(pk_tc, 2),
        "purchase_cost":  round(first_purch, 2),
        "leg_cost":       round(pk_tc + pk_toll + first_purch, 2),
        "cum_cost":       round(cum, 2),
        "tank_after_lt":  round(first_load_lt),
        "note":           f"Loaded {first_load_lt:,.0f} Lt  (capacity {cap_lt:,.0f} Lt)",
    })

    # ── Steps 2+: Deliver to each stop, reload if needed ─────────────────
    prev_lat, prev_lon = src_lat, src_lon
    stop_seq           = 0

    for i, stop in enumerate(ordered_stops):
        stop_seq += 1
        s_lat  = stop["station_lat"]
        s_lon  = stop["station_lon"]
        s_name = stop["station"]
        need   = stop["needed_lt"]
        need_mt= need / MT_TO_LITERS

        # ── RELOAD check: tank insufficient for this stop ─────────────────
        if tank_lt < need - 0.01:   # tiny tolerance for float precision
            n_reloads += 1

            # Back to source
            back_dist, back_toll = get_road_info(prev_lat, prev_lon, src_lat, src_lon)

            # Reload: enough for this and remaining stops
            remaining_needed = sum(s["needed_lt"] for s in ordered_stops[i:])
            reload_lt        = min(remaining_needed, cap_lt)
            reload_mt        = reload_lt / MT_TO_LITERS

            back_tc          = transport_cost_calc(back_dist, reload_mt)
            reload_purchase  = reload_mt * price_mt
            tot_transport   += back_tc
            tot_toll        += back_toll
            tot_purchase    += reload_purchase
            cum             += back_tc + back_toll + reload_purchase
            tank_lt          = reload_lt

            steps.append({
                "step_type":      "RELOAD",
                "label":          f"🔄 RELOAD at Source (trip {n_reloads})",
                "location":       f"{src_name}  ({src_id})",
                "qty_lt":         reload_lt,
                "qty_mt":         round(reload_mt, 3),
                "dist_km":        back_dist,
                "toll":           back_toll,
                "transport_cost": round(back_tc, 2),
                "purchase_cost":  round(reload_purchase, 2),
                "leg_cost":       round(back_tc + back_toll + reload_purchase, 2),
                "cum_cost":       round(cum, 2),
                "tank_after_lt":  round(reload_lt),
                "note":           (f"⚠ Tank was {tank_lt - reload_lt + reload_lt:,.0f} Lt, "
                                   f"needed {need:,.0f} Lt  |  "
                                   f"Back: {back_dist} km  |  "
                                   f"Reload: {reload_lt:,.0f} Lt  |  "
                                   f"Extra purchase: ₹{reload_purchase:,.0f}"),
            })
            prev_lat, prev_lon = src_lat, src_lon

        # ── Delivery leg ──────────────────────────────────────────────────
        del_dist, del_toll = get_road_info(prev_lat, prev_lon, s_lat, s_lon)

        del_tc        = transport_cost_calc(del_dist, need_mt)
        tot_transport += del_tc
        tot_toll      += del_toll
        # Purchase cost for this stop was already paid in LOAD/RELOAD step
        stop_purchase = need_mt * price_mt   # shown for reference only
        leg_c         = del_tc + del_toll
        cum          += leg_c + stop_purchase  # add purchase reference to cumulative
        tank_lt      -= need

        steps.append({
            "step_type":      "DELIVER",
            "label":          f"📍 STOP {stop_seq} – DELIVER",
            "location":       s_name,
            "qty_lt":         need,
            "qty_mt":         round(need_mt, 3),
            "dist_km":        del_dist,
            "toll":           del_toll,
            "transport_cost": round(del_tc, 2),
            "purchase_cost":  round(stop_purchase, 2),
            "leg_cost":       round(leg_c + stop_purchase, 2),
            "cum_cost":       round(cum, 2),
            "tank_after_lt":  round(max(tank_lt, 0)),
            "note":           f"Delivered {need:,.0f} Lt  |  Tank remaining: {max(tank_lt,0):,.0f} Lt",
        })

        prev_lat, prev_lon = s_lat, s_lon

    # ── Final PARK ────────────────────────────────────────────────────────
    final_station = ordered_stops[-1]["station"]
    steps.append({
        "step_type":      "FINAL_PARK",
        "label":          "🏁 FINAL PARK",
        "location":       final_station,
        "qty_lt":         None, "qty_mt":         None,
        "dist_km":        None, "toll":            None,
        "transport_cost": None, "purchase_cost":   None,
        "leg_cost":       None, "cum_cost":        None,
        "tank_after_lt":  0,
        "note":           "Truck parked here – next run starts from this location",
    })

    cost_summary = {
        "total_lt":       total_lt,
        "total_mt":       round(total_lt / MT_TO_LITERS, 3),
        "tot_purchase":   round(tot_purchase,  2),
        "tot_transport":  round(tot_transport, 2),
        "tot_toll":       round(tot_toll,      2),
        "grand_total":    round(tot_purchase + tot_transport + tot_toll, 2),
        "n_reloads":      n_reloads,
        "pk_src_dist":    pk_dist,
        "pk_src_toll":    pk_toll,
        "first_load_lt":  first_load_lt,
    }

    return steps, cost_summary

# ═══════════════════════════════════════════════════════════════════════════
#  CORE OPTIMIZATION
# ═══════════════════════════════════════════════════════════════════════════

def run_optimization():
    print("="*65)
    print("  LPG AUTO DISPATCH OPTIMIZER")
    print(f"  Date : {datetime.today().strftime('%d-%m-%Y  %H:%M')}")
    print("="*65)

    print("\n[1/6] Loading data …")
    stations, sources = load_data()
    print(f"      Stations: {len(stations)}   Sources: {len(sources)}")

    needing = stations[stations["Now"].str.strip().str.upper() == "NO"].copy()
    print(f"\n[2/6] Stations needing LPG: {len(needing)}")
    if needing.empty:
        print("      ✓ No deliveries needed today.")
        return None, None, None
    for _, r in needing.iterrows():
        print(f"      – {r['Stations']}  ({r['Usable Lt']:,.0f} Lt)")

    print("\n[3/6] Loading truck positions …")
    trucks = build_fleet(stations)
    start_pos = {t["truck_id"]: {"station": t["parked_station"],
                                  "lat": t["parked_lat"], "lon": t["parked_lon"]}
                 for t in trucks}
    for t in trucks[:5]:
        print(f"      {t['truck_id']} ({t['type']}) → {t['parked_station']}")
    if len(trucks) > 5:
        print(f"      … and {len(trucks)-5} more trucks")

    print("\n[4/6] Finding best source per station …")
    sources_list = [row for _, row in sources.iterrows()]
    needing_rows = [row for _, row in needing.iterrows()]

    # ── Build ALL (station, source) tasks upfront as flat list ──────────────
    # Each task is one API call: (station_idx, source_row, slat, slon, needed_mt)
    # One flat ThreadPoolExecutor fires all tasks at once – no nested pools.
    tasks = []
    for si, srow in enumerate(needing_rows):
        slat      = float(srow["lat"])
        slon      = float(srow["lon"])
        needed_mt = float(srow["Usable Lt"]) / MT_TO_LITERS
        # Pre-filter: top SOURCE_PREFETCH_N sources by APPROXIMATE TOTAL COST
        # (purchase + transport using haversine×1.3 distance)
        # This ensures the cheapest source is never filtered out.
        # Pure distance ranking misses sources that are slightly farther but much cheaper.
        def _approx_total_cost(s):
            d = haversine(slat, slon, float(s["lat"]), float(s["lon"])) * 1.3
            tc = transport_cost_calc(d, needed_mt)
            pc = float(s["Price / MT Ex Terminal"]) * needed_mt
            return tc + pc
        ranked = sorted(sources_list, key=_approx_total_cost)[:SOURCE_PREFETCH_N]
        for src in ranked:
            tasks.append((si, src, slat, slon, needed_mt))

    # Count cache hits to report
    hits   = sum(1 for _,src,slat,slon,_ in tasks
                 if _rkey(float(src["lat"]),float(src["lon"]),slat,slon) in _route_cache)
    misses = len(tasks) - hits
    print(f"      {len(needing_rows)} stations × top-{SOURCE_PREFETCH_N} sources = "
          f"{len(tasks)} route lookups  ({hits} cached, {misses} live API)")

    # ── Fire ALL lookups in one flat pool ────────────────────────────────────
    def _fetch_task(task):
        si, src, slat, slon, needed_mt = task
        d, t = get_road_info(float(src["lat"]), float(src["lon"]), slat, slon)
        tc   = transport_cost_calc(d, needed_mt)
        pc   = float(src["Price / MT Ex Terminal"]) * needed_mt
        return si, src, d, t, tc, pc, tc + pc + t

    # results_by_station[si] = list of (src, dist, toll, tc, pc, total)
    results_by_station = [[] for _ in needing_rows]
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        for si, src, d, t, tc, pc, total in ex.map(_fetch_task, tasks):
            results_by_station[si].append((src, d, t, tc, pc, total))

    # ── Pick best source per station ────────────────────────────────────────
    station_data = []
    for si, srow in enumerate(needing_rows):
        slat      = float(srow["lat"])
        slon      = float(srow["lon"])
        sname     = srow["Stations"]
        needed_lt = float(srow["Usable Lt"])
        needed_mt = needed_lt / MT_TO_LITERS
        best = min(results_by_station[si], key=lambda x: x[5])
        best_src, best_dist, best_toll, best_tc, best_pc, best_total = best
        station_data.append({
            "station":        sname,
            "station_lat":    slat,  "station_lon":    slon,
            "needed_lt":      needed_lt, "needed_mt":  needed_mt,
            "source_id":      best_src["Source_ID"],
            "source_name":    best_src["Source_Name"],
            "source_lat":     float(best_src["lat"]),
            "source_lon":     float(best_src["lon"]),
            "price_mt":       float(best_src["Price / MT Ex Terminal"]),
            "dist_km":        best_dist,
            "toll_cost":      best_toll,
            "transport_cost": best_tc,
            "purchase_cost":  best_pc,
            "total_cost":     best_total,
        })
        print(f"      ✓ {sname[:45]:<45} → {best_src['Source_ID']}  ₹{best_total:,.0f}")

    print(f"\n[5/6] Grouping, assigning trucks & building journeys …")

    # ── Smart prewarm: only fetch pairs actually needed by Step 5 ───────────
    # Compute exactly which legs build_journey will call, NOT all combinations.
    def _prewarm_exact(station_data_local, trucks_local):
        pairs = set()

        # Group stations by source (mirrors Step 5 logic)
        by_src = {}
        for sd in station_data_local:
            by_src.setdefault(sd["source_id"], []).append(sd)

        for src_id, sds in by_src.items():
            src_lat = sds[0]["source_lat"]
            src_lon = sds[0]["source_lon"]

            # Sort and partition exactly as Step 5 will
            for sd in sds:
                sd["_d_src"] = haversine(src_lat, src_lon,
                                          sd["station_lat"], sd["station_lon"])
            sds_sorted = sorted(sds, key=lambda x: x["_d_src"])
            runs = balanced_partition(sds_sorted, MAX_STOPS_PER_TRUCK, MAX_GROUPING_KM)

            for run in runs:
                total_lt = sum(r["needed_lt"] for r in run)

                # All truck→source legs (find nearest trucks by haversine)
                nearest_trucks = sorted(
                    trucks_local,
                    key=lambda t: haversine(t["parked_lat"],t["parked_lon"],src_lat,src_lon)
                )[:3]   # only top-3 nearest trucks will realistically be chosen
                for t in nearest_trucks:
                    pairs.add((t["parked_lat"], t["parked_lon"], src_lat, src_lon))

                # Source → stop1 → stop2 → stop3 (exact sequence)
                prev_lat, prev_lon = src_lat, src_lon
                for stop in run:
                    pairs.add((prev_lat, prev_lon,
                               stop["station_lat"], stop["station_lon"]))
                    # Potential reload back-leg
                    pairs.add((stop["station_lat"], stop["station_lon"],
                               src_lat, src_lon))
                    prev_lat = stop["station_lat"]
                    prev_lon = stop["station_lon"]

        uncached = [p for p in pairs if _rkey(p[0],p[1],p[2],p[3]) not in _route_cache]
        if uncached:
            print(f"      Pre-warming {len(uncached)} exact route pairs "
                  f"({len(pairs)-len(uncached)} cached) …")
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
                list(ex.map(lambda p: get_road_info(p[0],p[1],p[2],p[3]), uncached))
        else:
            print(f"      All {len(pairs)} route pairs already cached")

    _prewarm_exact(station_data, trucks)

    by_source = defaultdict(list)
    for sd in station_data:
        by_source[sd["source_id"]].append(sd)

    delivery_plans  = []
    truck_available = {t["truck_id"]: True for t in trucks}
    truck_by_id     = {t["truck_id"]: t    for t in trucks}

    for src_id, sds in by_source.items():
        src_lat  = sds[0]["source_lat"];  src_lon  = sds[0]["source_lon"]
        src_name = sds[0]["source_name"]; price_mt = sds[0]["price_mt"]

        for sd in sds:
            sd["_d_src"] = haversine(src_lat, src_lon, sd["station_lat"], sd["station_lon"])
        sds.sort(key=lambda x: x["_d_src"])
        runs = balanced_partition(sds, MAX_STOPS_PER_TRUCK, MAX_GROUPING_KM)

        for run in runs:
            total_needed_lt = sum(r["needed_lt"] for r in run)

            # Prefer truck that fits everything (no reload); else pick nearest available
            cands_all = [(haversine(t["parked_lat"],t["parked_lon"],src_lat,src_lon), t)
                         for t in trucks if truck_available[t["truck_id"]]]
            if not cands_all:
                # All trucks deployed – flag unserved stations and skip this run
                unserved_names = [s["station"] for s in run]
                print(f"      ❌ NO TRUCK AVAILABLE for: {unserved_names}")
                print(f"         These stations will be added to UNSERVED list.")
                for s in run:
                    _unserved_stations.append(s)
                continue
            fitting  = sorted([(d,t) for d,t in cands_all if t["capacity_lt"] >= total_needed_lt],
                               key=lambda x: x[0])
            cands    = fitting if fitting else sorted(cands_all, key=lambda x: x[0])
            _, chosen = cands[0]
            truck_available[chosen["truck_id"]] = False

            # Build full journey with reload detection
            journey_steps, costs = build_journey(
                chosen, start_pos[chosen["truck_id"]],
                src_id, src_name, src_lat, src_lon, run, price_mt)

            final_park = run[-1]["station"]
            final_lat  = run[-1]["station_lat"]
            final_lon  = run[-1]["station_lon"]

            reload_flag = f"  ⚠ {costs['n_reloads']} reload" if costs["n_reloads"] else ""
            print(f"      {chosen['truck_id']} ({chosen['type']}) | {src_id} | "
                  f"[{start_pos[chosen['truck_id']]['station'][:20]}] → "
                  f"{' → '.join(s['station'][:16] for s in run)}{reload_flag}")

            delivery_plans.append({
                "truck_id":       chosen["truck_id"],
                "truck_type":     chosen["type"],
                "capacity_lt":    chosen["capacity_lt"],
                "initial_park":   start_pos[chosen["truck_id"]]["station"],
                "source_id":      src_id,
                "source_name":    src_name,
                "source_lat":     src_lat, "source_lon": src_lon,
                "pk_src_dist":    costs["pk_src_dist"],
                "pk_src_toll":    costs["pk_src_toll"],
                "first_load_lt":  costs["first_load_lt"],
                "stops":          run,
                "journey_steps":  journey_steps,
                "final_park":     final_park,
                "final_lat":      final_lat, "final_lon": final_lon,
                "total_lt":       costs["total_lt"],
                "total_mt":       costs["total_mt"],
                "tot_purchase":   costs["tot_purchase"],
                "tot_transport":  costs["tot_transport"],
                "tot_toll":       costs["tot_toll"],
                "grand_total":    costs["grand_total"],
                "n_reloads":      costs["n_reloads"],
            })

            truck_by_id[chosen["truck_id"]]["parked_station"] = final_park
            truck_by_id[chosen["truck_id"]]["parked_lat"]     = final_lat
            truck_by_id[chosen["truck_id"]]["parked_lon"]     = final_lon

    print(f"\n      Total runs   : {len(delivery_plans)}")
    if _unserved_stations:
        print(f"\n      ❌ UNSERVED STATIONS ({len(_unserved_stations)}) – no truck available:")
        for s in _unserved_stations:
            print(f"         – {s['station']}  ({s['needed_lt']:,.0f} Lt)")
    total_reloads = sum(dp["n_reloads"] for dp in delivery_plans)
    if total_reloads:
        print(f"      ⚠ Reload trips : {total_reloads}  (extra cost included in totals)")

    used_ids = {dp["truck_id"] for dp in delivery_plans}
    fleet_status = [{"truck_id": t["truck_id"], "type": t["type"],
                     "status": "DEPLOYED" if t["truck_id"] in used_ids else "STANDBY",
                     "initial_park": start_pos[t["truck_id"]]["station"],
                     "final_park":   truck_by_id[t["truck_id"]]["parked_station"]}
                    for t in trucks]

    print("\n      Saving truck positions …")
    save_positions(trucks)

    # ── Resolve unserved stations ─────────────────────────────────────────
    unserved_resolutions = []
    if _unserved_stations:
        print(f"\n[6a] Resolving {len(_unserved_stations)} unserved stations …")
        try:
            stations_df, sources_df = load_data()
        except Exception:
            stations_df, sources_df = None, None
        unserved_resolutions = resolve_unserved(
            _unserved_stations, delivery_plans, stations_df, sources_df)
        for r in unserved_resolutions:
            print(f"      {r['station'][:40]:<40} → {r['when']} | {r['action']}")
            if r.get('swap_candidate'):
                print(f"         💡 Swap: {r['swap_candidate'][:80]}")

    return delivery_plans, fleet_status, trucks, _unserved_stations, unserved_resolutions


# ═══════════════════════════════════════════════════════════════════════════
#  PRIORITY SCORING & UNSERVED STATION RESOLUTION
# ═══════════════════════════════════════════════════════════════════════════

def priority_score(capacity_lt, usable_lt):
    """
    Score 0-100 for how urgently a station needs LPG.
    urgency (60%) = usable_lt / capacity (higher usable = tank is emptier = more urgent)
    volume  (40%) = usable_lt / max_truck_capacity (larger delivery = more people affected)
    """
    max_cap = 12 * MT_TO_LITERS  # 21720 Lt
    urgency = min(usable_lt / capacity_lt, 1.0) if capacity_lt > 0 else 0.5
    volume  = min(usable_lt / max_cap, 1.0)
    return round((urgency * 0.6 + volume * 0.4) * 100, 1)

def estimate_run_duration_min(journey_steps):
    """
    Estimate total time (minutes) for a truck run based on journey steps.
    Uses AVG_SPEED_KMH for driving legs, LOAD/UNLOAD constants for stops.
    """
    total_min = 0.0
    for step in journey_steps:
        stype = step.get("step_type", "")
        dist  = step.get("dist_km") or 0

        if stype in ("LOAD", "RELOAD"):
            # Drive to source + load time
            total_min += (dist / AVG_SPEED_KMH) * 60 + LOAD_MIN_AT_SOURCE
        elif stype == "DELIVER":
            # Drive to station + unload time
            total_min += (dist / AVG_SPEED_KMH) * 60 + UNLOAD_MIN_PER_STOP

    return round(total_min, 1)

def resolve_unserved(unserved_stations, delivery_plans, all_stations_df, sources_df):
    """
    For each unserved station, determine the best resolution.
    Handles 5 edge cases:
      1. Capacity check   – truck must fit the unserved station's LPG need
      2. Deduplication    – each truck assigned to at most one unserved station today
      3. Grouping         – nearby unserved stations batched into same truck run
      4. Oversize warning – stations needing > max truck capacity flagged
      5. Swap validation  – only suggest swapping stops without reloads (simple runs)
    """
    resolutions = []

    # Build priority map for all stations
    station_priority = {}
    if all_stations_df is not None:
        for _, row in all_stations_df.iterrows():
            name = str(row.get("Stations", "")).strip()
            cap  = float(row.get("Capacity in Lt", 0) or 0)
            usl  = float(row.get("Usable Lt", 0) or 0)
            if cap > 0 and usl > 0:
                station_priority[name] = priority_score(cap, usl)

    # Build source lookup
    source_by_id = {}
    if sources_df is not None:
        for _, s in sources_df.iterrows():
            source_by_id[s["Source_ID"]] = s

    max_truck_cap = max(ft["capacity_lt"] for ft in FLEET)  # 21720 Lt

    # ── EDGE CASE 3: Group nearby unserved stations together ─────────────────
    # Stations within MAX_GROUPING_KM of each other can share a truck run.
    remaining_unserved = list(unserved_stations)
    groups = []
    while remaining_unserved:
        base  = remaining_unserved.pop(0)
        group = [base]
        still = []
        for s in remaining_unserved:
            if haversine(base["station_lat"], base["station_lon"],
                         s["station_lat"],   s["station_lon"]) <= MAX_GROUPING_KM:
                group.append(s)
            else:
                still.append(s)
        remaining_unserved = still
        groups.append(group)

    # Track which trucks have already been assigned a TODAY extra run (dedup)
    trucks_assigned_today = set()

    for group in groups:
        group_lt    = sum(s["needed_lt"] for s in group)
        group_names = [s["station"] for s in group]

        # Representative station for distance calculations (centroid)
        centroid_lat = sum(s["station_lat"] for s in group) / len(group)
        centroid_lon = sum(s["station_lon"] for s in group) / len(group)

        # ── EDGE CASE 4: Oversize warning ─────────────────────────────────────
        oversize_note = ""
        if group_lt > max_truck_cap:
            n_loads = -(-group_lt // max_truck_cap)  # ceiling division
            oversize_note = (f" ⚠ Needs {group_lt:,} Lt > max truck {max_truck_cap:,} Lt. "
                             f"Requires {n_loads} loads / reload trips.")

        # Build one resolution record per station in group
        # (all stations in a group share the same truck assignment)
        group_resolution = {
            "grouped_with": group_names if len(group) > 1 else None,
            "group_lt":     group_lt,
            "oversize":     oversize_note,
        }

        # ── PHASE 1: Can any truck serve this group TODAY? ────────────────────
        best_truck  = None
        best_remain = -1
        best_finish = None
        best_time_needed = None

        for dp in delivery_plans:
            tid = dp["truck_id"]

            # EDGE CASE 2: Skip trucks already assigned an extra run today
            if tid in trucks_assigned_today:
                continue

            est_dur   = estimate_run_duration_min(dp.get("journey_steps", []))
            remaining = WORK_DAY_MIN - est_dur
            if remaining <= 0:
                continue

            # EDGE CASE 1: Capacity check — truck must carry the group's LPG
            truck_cap = dp.get("capacity_lt", max_truck_cap)
            if truck_cap < min(s["needed_lt"] for s in group):
                # Can't even do one stop — skip
                continue
            # Note: if group_lt > truck_cap, reloads will be needed but it's still feasible

            # Find nearest source to centroid
            if source_by_id:
                all_dists = [(sid, haversine(centroid_lat, centroid_lon,
                                              float(s["lat"]), float(s["lon"])) * 1.3)
                             for sid, s in source_by_id.items()]
                _, nearest_src_dist = min(all_dists, key=lambda x: x[1])
            else:
                nearest_src_dist = haversine(centroid_lat, centroid_lon,
                                              dp["source_lat"], dp["source_lon"]) * 1.3

            park_lat = dp["final_lat"]
            park_lon = dp["final_lon"]
            park_to_centroid = haversine(park_lat, park_lon,
                                          centroid_lat, centroid_lon) * 1.3
            n_stops = len(group)
            # Estimate extra run time (may include reloads if group_lt > truck_cap)
            n_reloads_extra = max(0, -(-group_lt // truck_cap) - 1)
            time_needed = (
                (park_to_centroid / AVG_SPEED_KMH) * 60   # park → centroid area
                + LOAD_MIN_AT_SOURCE                        # load at source
                + (nearest_src_dist / AVG_SPEED_KMH) * 60  # source → first stop
                + n_stops * UNLOAD_MIN_PER_STOP             # unloading
                + n_stops * (15 / AVG_SPEED_KMH) * 60      # inter-stop driving (est 15km)
                + n_reloads_extra * LOAD_MIN_AT_SOURCE      # reload time if needed
            )

            if remaining >= time_needed and remaining > best_remain:
                best_remain      = remaining
                best_truck       = dp
                best_finish      = est_dur
                best_time_needed = time_needed

        # ── PHASE 2: Tomorrow scheduling ─────────────────────────────────────
        if best_truck is None and delivery_plans:
            # Find truck whose final park is nearest to the group centroid
            best_truck = min(
                delivery_plans,
                key=lambda dp: haversine(dp["final_lat"], dp["final_lon"],
                                          centroid_lat, centroid_lon)
            )
            best_finish = estimate_run_duration_min(best_truck.get("journey_steps", []))

        # ── PHASE 3: Swap suggestion ──────────────────────────────────────────
        # EDGE CASE 5: Only suggest swapping stops from SIMPLE runs (no reloads)
        best_swap = None
        for u in group:
            u_score = station_priority.get(u["station"], 50.0)
            for dp in delivery_plans:
                if dp.get("n_reloads", 0) > 0:
                    continue   # don't suggest swapping from reload runs – too complex
                for stop in dp.get("stops", []):
                    s_score = station_priority.get(stop["station"], 50.0)
                    if s_score < u_score:
                        truck_cap = dp.get("capacity_lt", max_truck_cap)
                        if u["needed_lt"] <= truck_cap:
                            best_swap = (
                                f"Drop '{stop['station']}' (score {s_score:.0f}) "
                                f"from {dp['truck_id']} (no-reload run) → serve "
                                f"'{u['station']}' (score {u_score:.0f}) instead. "
                                f"Both need LPG but '{u['station']}' is more urgent.")
                        break
                if best_swap:
                    break
            if best_swap:
                break

        # ── Build one record per station in this group ────────────────────────
        for u in group:
            u_score = station_priority.get(u["station"], 50.0)

            if best_truck and best_remain > 0:   # TODAY
                trucks_assigned_today.add(best_truck["truck_id"])
                group_note = (f" | Grouped with: {[s['station'] for s in group if s is not u]}"
                              if len(group) > 1 else "")
                action        = "REASSIGN TODAY"
                action_detail = (
                    f"Truck {best_truck['truck_id']} finishes in ~{best_finish:.0f} min "
                    f"({best_finish/60:.1f}h), has {best_remain:.0f} min left. "
                    f"Est extra run: {best_time_needed:.0f} min.{group_note}{oversize_note}")
                when          = "TODAY"
                remaining_out = round(best_remain, 0)
                time_out      = round(best_time_needed, 0)

            elif best_truck:                      # TOMORROW
                dist_km = round(haversine(best_truck["final_lat"], best_truck["final_lon"],
                                           u["station_lat"], u["station_lon"]) * 1.3, 1)
                group_note = (f" | Grouped with: {[s['station'] for s in group if s is not u]}"
                              if len(group) > 1 else "")
                action        = "SCHEDULE TOMORROW"
                action_detail = (
                    f"Truck {best_truck['truck_id']} parks {dist_km} km away. "
                    f"First assignment tomorrow morning.{group_note}{oversize_note}")
                when          = "TOMORROW"
                remaining_out = round(max(0, WORK_DAY_MIN - best_finish), 0)
                time_out      = None

            else:
                action        = "MANUAL REVIEW"
                action_detail = f"No trucks available. Review fleet manually.{oversize_note}"
                when          = "TBD"
                remaining_out = None
                time_out      = None

            resolutions.append({
                "station":          u["station"],
                "needed_lt":        u["needed_lt"],
                "needed_mt":        round(u["needed_mt"], 3),
                "priority_score":   u_score,
                "action":           action,
                "action_detail":    action_detail,
                "truck_id":         best_truck["truck_id"] if best_truck else "–",
                "truck_type":       best_truck.get("truck_type","–") if best_truck else "–",
                "est_finish_min":   round(best_finish, 0) if best_finish else None,
                "remaining_min":    remaining_out,
                "time_needed_min":  time_out,
                "when":             when,
                "swap_candidate":   best_swap,
            })

    return resolutions

# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL STYLES
# ═══════════════════════════════════════════════════════════════════════════

C = {
    "BD":"1F4E79","BL":"D6E4F0","BA":"EBF3FB",
    "GD":"375623","GL":"E2EFDA",
    "OD":"C55A11","OL":"FCE4D6",
    "PD":"6B2C91","YL":"FFF2CC",
    "GR":"F2F2F2","WH":"FFFFFF",
    "DEP":"C6EFCE","SBY":"FFEB9C",
    # Step-type colours
    "INITIAL": "D6E4F0",   # light blue
    "SOURCE":  "FFF2CC",   # yellow
    "DELIVER": "E2EFDA",   # green
    "RELOAD":  "FFD0D0",   # light red – stands out clearly
    "FINAL":   "FCE4D6",   # orange
}
MF="#,##0.00"; NF="#,##0"

STEP_FILL = {
    "INITIAL_PARK": C["INITIAL"],
    "LOAD":         C["SOURCE"],
    "DELIVER":      C["DELIVER"],
    "RELOAD":       C["RELOAD"],
    "FINAL_PARK":   C["FINAL"],
}

def _f(h):  return PatternFill("solid",start_color=h)
def _fn(bold=False,sz=10,col="000000"):
    return Font(name="Arial",bold=bold,size=sz,color=col)
def _bd():
    s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)
def _ac(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def _al(): return Alignment(horizontal="left",  vertical="center",wrap_text=True)

def hdr(ws,r,c,v,fh="1F4E79"):
    x=ws.cell(row=r,column=c,value=v)
    x.font=Font(name="Arial",bold=True,size=10,color="FFFFFF")
    x.fill=_f(fh); x.border=_bd(); x.alignment=_ac()

def cel(ws,r,c,v,fmt=None,fh=None,bold=False,center=False):
    x=ws.cell(row=r,column=c,value=v)
    x.font=_fn(bold=bold); x.border=_bd()
    x.alignment=_ac() if center else _al()
    if fmt: x.number_format=fmt
    if fh:  x.fill=_f(fh)

def stitle(ws,rng,text,fh="1F4E79"):
    ws.merge_cells(rng)
    x=ws[rng.split(":")[0]]
    x.value=text; x.fill=_f(fh)
    x.font=Font(name="Arial",bold=True,size=13,color="FFFFFF")
    x.alignment=_ac(); ws.row_dimensions[1].height=30

def cw(ws,wl):
    for i,w in enumerate(wl,1):
        ws.column_dimensions[get_column_letter(i)].width=w

# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════════════

def build_excel(delivery_plans, fleet_status, trucks, unserved, unserved_resolutions, output_path):
    wb    = Workbook()
    today = datetime.today().strftime("%d-%m-%Y")

    # ── Sheet 1: Delivery Plan Summary ─────────────────────────────────────
    ws1=wb.active; ws1.title="Delivery Plan"; ws1.freeze_panes="A3"
    stitle(ws1,"A1:S1",f"LPG AUTO DISPATCH – DELIVERY PLAN   ({today})")
    h1=["Truck\nID","Type","Cap\n(Lt)","Starts From",
        "Source\nID","Source Name",
        "Truck→Src\n(km)","Truck→Src\nToll (₹)","First\nLoad (Lt)","Reloads",
        "Stop 1","Stop 2","Stop 3",
        "Total\n(Lt)","Total\n(MT)",
        "Purchase\n(₹)","Transport\n(₹)","Toll\n(₹)","Grand Total\n(₹)"]
    [hdr(ws1,2,i+1,h) for i,h in enumerate(h1)]
    ws1.row_dimensions[2].height=44

    for ri,dp in enumerate(delivery_plans,3):
        alt=C["BA"] if ri%2==0 else C["WH"]
        sn=[s["station"] for s in dp["stops"]]
        while len(sn)<3: sn.append("")
        reload_disp = f"⚠ {dp['n_reloads']}" if dp["n_reloads"] else "–"
        vals=[dp["truck_id"],dp["truck_type"],dp["capacity_lt"],dp["initial_park"],
              dp["source_id"],dp["source_name"],
              dp["pk_src_dist"],dp["pk_src_toll"],dp["first_load_lt"],reload_disp,
              sn[0],sn[1],sn[2],
              dp["total_lt"],dp["total_mt"],
              dp["tot_purchase"],dp["tot_transport"],dp["tot_toll"],dp["grand_total"]]
        fmts=[None,None,NF,None,None,None,"0.1",MF,NF,None,
              None,None,None,NF,"0.000",MF,MF,MF,MF]
        for ci,(v,f) in enumerate(zip(vals,fmts),1):
            rfh = C["RELOAD"] if (ci==10 and dp["n_reloads"]>0) else alt
            cel(ws1,ri,ci,v,fmt=f,fh=rfh,bold=(ci==10 and dp["n_reloads"]>0))

    tr=len(delivery_plans)+3
    ws1.cell(row=tr,column=1,value="TOTALS").font=_fn(bold=True)
    for col,fmt in [(14,NF),(15,"0.000"),(16,MF),(17,MF),(18,MF),(19,MF)]:
        c=ws1.cell(row=tr,column=col)
        c.value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{tr-1})"
        c.number_format=fmt; c.font=_fn(bold=True)
        c.border=_bd(); c.fill=_f(C["BL"]); c.alignment=_ac()
    cw(ws1,[9,9,12,30,10,24,12,14,12,10,28,28,28,13,13,18,18,13,20])

    # ── Sheet 2: Full Journey Step-by-Step ─────────────────────────────────
    # Every step for every truck including RELOAD legs.
    ws2=wb.create_sheet("Delivery Sequence"); ws2.freeze_panes="A3"
    stitle(ws2,"A1:O1",
           f"FULL JOURNEY – STEP BY STEP  (incl. RELOAD)   ({today})",C["GD"])

    h2=["Truck\nID","Type","Step\n#","Step Type","Location",
        "Qty\n(Lt)","Qty\n(MT)","Dist\n(km)","Toll\n(₹)",
        "Transport\n(₹)","Purchase\n(₹)",
        "Leg Cost\n(₹)","Cumulative\n(₹)",
        "Tank\nAfter (Lt)","Notes"]
    [hdr(ws2,2,i+1,h,fh=C["GD"]) for i,h in enumerate(h2)]
    ws2.row_dimensions[2].height=44

    # Colour legend in header row
    legend_items=[
        ("🚚 INITIAL PARK",C["INITIAL"]),
        ("⛽ SOURCE / LOAD",C["SOURCE"]),
        ("📍 DELIVER",C["DELIVER"]),
        ("🔄 RELOAD ← extra cost",C["RELOAD"]),
        ("🏁 FINAL PARK",C["FINAL"]),
    ]
    # place legend to the right of headers (cols 16–20)
    for li,(txt,fh) in enumerate(legend_items):
        c=ws2.cell(row=2,column=16+li,value=txt)
        c.fill=_f(fh); c.font=_fn(bold=True,sz=9)
        c.border=_bd(); c.alignment=_ac()
        ws2.column_dimensions[get_column_letter(16+li)].width=22

    r2=3
    for dp in delivery_plans:
        for si, step in enumerate(dp["journey_steps"], 1):
            fh      = STEP_FILL.get(step["step_type"], C["WH"])
            is_rel  = step["step_type"] == "RELOAD"
            is_init = step["step_type"] == "INITIAL_PARK"
            is_fin  = step["step_type"] == "FINAL_PARK"

            vals=[
                dp["truck_id"], dp["truck_type"], si,
                step["label"], step["location"],
                step["qty_lt"], step["qty_mt"],
                step["dist_km"], step["toll"],
                step["transport_cost"], step["purchase_cost"],
                step["leg_cost"], step["cum_cost"],
                step["tank_after_lt"], step["note"],
            ]
            fmts=[None,None,None,None,None,
                  NF,"0.000","0.1",MF,MF,MF,MF,MF,NF,None]

            for ci,(v,f) in enumerate(zip(vals,fmts),1):
                bold = is_rel and ci in [4,5,15]
                cel(ws2,r2,ci,v,fmt=f,fh=fh,bold=bold)

            # Highlight the entire RELOAD row with red font too
            if is_rel:
                for ci in range(1,16):
                    ws2.cell(row=r2,column=ci).font = _fn(bold=(ci in [4,5,15]),
                                                          col="AA0000")
            r2+=1
        r2+=1   # blank separator between trucks

    cw(ws2,[9,9,6,28,40,12,11,10,12,14,14,14,14,14,45])

    # ── Sheet 3: Fleet Status ────────────────────────────────────────────────
    ws3=wb.create_sheet("Fleet Status"); ws3.freeze_panes="A3"
    stitle(ws3,"A1:F1",f"FLEET END-OF-DAY STATUS   ({today})",C["PD"])
    h3=["Truck\nID","Type","Status","Initial Parked At","Final Parked At","Notes"]
    [hdr(ws3,2,i+1,h,fh=C["PD"]) for i,h in enumerate(h3)]
    ws3.row_dimensions[2].height=44
    for ri,ts in enumerate(fleet_status,3):
        alt=C["GR"] if ri%2==0 else C["WH"]
        sf=C["DEP"] if ts["status"]=="DEPLOYED" else C["SBY"]
        note="✅ Delivered today" if ts["status"]=="DEPLOYED" else "– On standby"
        for ci,v in enumerate([ts["truck_id"],ts["type"],ts["status"],
                                 ts["initial_park"],ts["final_park"],note],1):
            cel(ws3,ri,ci,v,fh=(sf if ci==3 else alt))
    last=len(fleet_status)+4
    dn=sum(1 for t in fleet_status if t["status"]=="DEPLOYED")
    for lbl,val,row in [("DEPLOYED",dn,last),("STANDBY",len(fleet_status)-dn,last+1)]:
        ws3.cell(row=row,column=1,value=lbl).font=_fn(bold=True)
        ws3.cell(row=row,column=2,value=val).font=_fn(bold=True)
    cw(ws3,[9,10,12,44,44,30])

    # ── Sheet 4: Cost Summary ────────────────────────────────────────────────
    ws4=wb.create_sheet("Cost Summary"); ws4.freeze_panes="A3"
    stitle(ws4,"A1:H1",f"COST SUMMARY   ({today})",C["OD"])
    h4=["Truck\nID","Source\nID","Stations Served",
        "Purchase\n(₹)","Transport\n(₹)","Toll\n(₹)","Grand Total\n(₹)","Reloads"]
    [hdr(ws4,2,i+1,h,fh=C["OD"]) for i,h in enumerate(h4)]
    ws4.row_dimensions[2].height=44
    for ri,dp in enumerate(delivery_plans,3):
        alt=C["OL"] if ri%2==0 else C["WH"]
        seq=" → ".join(s["station"] for s in dp["stops"])
        rdsp = (f"⚠ {dp['n_reloads']} reload{'s' if dp['n_reloads']>1 else ''}"
                if dp["n_reloads"] else "–")
        for ci,(v,f) in enumerate(zip(
            [dp["truck_id"],dp["source_id"],seq,
             dp["tot_purchase"],dp["tot_transport"],dp["tot_toll"],dp["grand_total"],rdsp],
            [None,None,None,MF,MF,MF,MF,None]),1):
            rfh = C["RELOAD"] if (ci==8 and dp["n_reloads"]>0) else alt
            cel(ws4,ri,ci,v,fmt=f,fh=rfh,bold=(ci==8 and dp["n_reloads"]>0))
    tr4=len(delivery_plans)+3
    ws4.cell(row=tr4,column=1,value="TOTALS").font=_fn(bold=True)
    for col,fmt in [(4,MF),(5,MF),(6,MF),(7,MF)]:
        c=ws4.cell(row=tr4,column=col)
        c.value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{tr4-1})"
        c.number_format=fmt; c.font=_fn(bold=True)
        c.border=_bd(); c.fill=_f(C["OL"]); c.alignment=_ac()
    cw(ws4,[9,10,60,20,20,16,22,18])

    # ── Sheet 5: Unserved Stations (only if any) ──────────────────────────
    if unserved:
        ws5=wb.create_sheet("Unserved Stations"); ws5.freeze_panes="A3"
        stitle(ws5,"A1:D1",f"UNSERVED STATIONS – NO TRUCK AVAILABLE   ({today})","AA0000")
        h5=["Station Name","Needed (Lt)","Needed (MT)","Reason"]
        [hdr(ws5,2,i+1,h,fh="AA0000") for i,h in enumerate(h5)]
        ws5.row_dimensions[2].height=40
        for ri,s in enumerate(unserved,3):
            alt=C["RELOAD"] if ri%2==0 else "FFE8E8"
            cel(ws5,ri,1,s["station"],fh=alt)
            cel(ws5,ri,2,s["needed_lt"],fmt=NF,fh=alt)
            cel(ws5,ri,3,round(s.get("needed_mt",s["needed_lt"]/1810),3),fmt="0.000",fh=alt)
            cel(ws5,ri,4,"All 30 trucks already deployed – increase fleet or reduce stations",fh=alt)
        cw(ws5,[44,14,14,55])
        print(f"      ⚠  Unserved stations sheet added ({len(unserved)} stations)")

    # ── Sheet 6: Unserved Station Resolution ────────────────────────────────
    if unserved_resolutions:
        ws6 = wb.create_sheet("Unserved – Resolution Plan"); ws6.freeze_panes="A3"
        stitle(ws6,"A1:I1",
               f"UNSERVED STATIONS – RESOLUTION PLAN   ({today})","AA0000")

        # Legend row
        leg_items = [("REASSIGN TODAY","C6EFCE"),("SCHEDULE TOMORROW","FFF2CC"),
                     ("SWAP SUGGESTED","FFD0D0"),("MANUAL REVIEW","F2F2F2")]
        for li,(txt,fh) in enumerate(leg_items):
            c = ws6.cell(row=2, column=6+li, value=txt)
            c.fill=_f(fh); c.font=_fn(bold=True,sz=9)
            c.border=_bd(); c.alignment=_ac()
            ws6.column_dimensions[get_column_letter(6+li)].width=22

        h6=["Station Name","Needed (Lt)","Needed (MT)","Priority Score",
            "Recommended Action","Truck","Est Finish (min)",
            "Time Left (min)","Time Needed (min)","When","Swap Suggestion"]
        [hdr(ws6,3,i+1,h,fh="AA0000") for i,h in enumerate(h6)]
        ws6.row_dimensions[3].height=44

        ACTION_COLOR = {
            "REASSIGN TODAY":   "C6EFCE",
            "SCHEDULE TOMORROW":"FFF2CC",
            "MANUAL REVIEW":    "F2F2F2",
        }

        for ri, r in enumerate(unserved_resolutions, 4):
            fh = ACTION_COLOR.get(r["action"], C["WH"])
            # If swap suggested, tint red regardless
            if r.get("swap_candidate"):
                fh = "FFD0D0"

            vals = [
                r["station"],
                r["needed_lt"],
                r["needed_mt"],
                r["priority_score"],
                r["action"],
                r["truck_id"] or "–",
                r["est_finish_min"] or "–",
                r["remaining_min"] or "–",
                r["time_needed_min"] or "–",
                r["when"] or "–",
                r.get("swap_candidate") or "–",
            ]
            fmts = [None,NF,"0.000","0.0",None,None,NF,NF,NF,None,None]
            for ci,(v,f) in enumerate(zip(vals,fmts),1):
                bold = (ci == 5 and r["action"] == "REASSIGN TODAY")
                cel(ws6, ri, ci, v, fmt=f, fh=fh, bold=bold)

        cw(ws6,[42,12,12,14,20,10,16,14,16,12,60])

        # Summary box
        sr = len(unserved_resolutions) + 6
        ws6.cell(row=sr,column=1,value="SUMMARY").font=_fn(bold=True,sz=11)
        today_count    = sum(1 for r in unserved_resolutions if r["when"]=="TODAY")
        tomorrow_count = sum(1 for r in unserved_resolutions if r["when"]=="TOMORROW")
        swap_count     = sum(1 for r in unserved_resolutions if r.get("swap_candidate"))
        for txt, val, row, fh in [
            ("Can serve TODAY (truck finishes early):", today_count,    sr+1, "C6EFCE"),
            ("Scheduled TOMORROW (truck reassigned):", tomorrow_count,  sr+2, "FFF2CC"),
            ("Swap suggestions (reprioritise run):",   swap_count,      sr+3, "FFD0D0"),
        ]:
            ws6.cell(row=row,column=1,value=txt).font=_fn(bold=True)
            c = ws6.cell(row=row,column=2,value=val)
            c.font=_fn(bold=True); c.fill=_f(fh); c.border=_bd(); c.alignment=_ac()

    wb.save(output_path)
    print(f"\n[6/6] Excel saved → {output_path}")
    print("="*65)

# ═══════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

if __name__=="__main__":
    delivery_plans, fleet_status, trucks, unserved, unserved_resolutions = run_optimization()
    if delivery_plans is None:
        sys.exit(0)

    date_str = datetime.today().strftime("%Y%m%d_%H%M")
    out_path = os.path.join(output_dir(), f"LPG_DISPATCH_PLAN_{date_str}.xlsx")
    build_excel(delivery_plans, fleet_status, trucks, unserved, unserved_resolutions, out_path)

    print("\n📋 QUICK SUMMARY")
    print(f"   Runs planned  : {len(delivery_plans)}")
    print(f"   Trucks used   : "
          f"{sum(1 for t in fleet_status if t['status']=='DEPLOYED')} / {len(trucks)}")
    print(f"   Total LPG     : {sum(dp['total_lt'] for dp in delivery_plans):,.0f} Lt")
    print(f"   Total cost    : ₹{sum(dp['grand_total'] for dp in delivery_plans):,.2f}")
    total_reloads = sum(dp["n_reloads"] for dp in delivery_plans)
    if total_reloads:
        print(f"   ⚠ Reload trips: {total_reloads}  (extra transport+purchase included)")
    if unserved:
        print(f"   ❌ Unserved     : {len(unserved)} stations")
        today_r    = sum(1 for r in unserved_resolutions if r["when"]=="TODAY")
        tomorrow_r = sum(1 for r in unserved_resolutions if r["when"]=="TOMORROW")
        swap_r     = sum(1 for r in unserved_resolutions if r.get("swap_candidate"))
        if today_r:    print(f"      ✓ {today_r} can be served TODAY (truck available)")
        if tomorrow_r: print(f"      📅 {tomorrow_r} scheduled TOMORROW")
        if swap_r:     print(f"      💡 {swap_r} swap suggestion(s) available")
    print(f"\n   Output        : {out_path}\n")